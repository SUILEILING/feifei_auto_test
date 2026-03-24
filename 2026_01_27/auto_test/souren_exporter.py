from lib.var import *

class ResultExporter:
    """结果导出器 - 将JSON结果自动转换为Excel,使用matplotlib绘制柱状图"""
    
    def __init__(self):
        # 延迟导入，避免循环导入问题
        import souren_config
        
        # 获取配置
        self.log_dir = souren_config.EXECUTION_DIR
        
        # 应用配置值（使用默认值避免缺失配置）
        self.default_row_height = getattr(souren_config, 'EXCEL_DEFAULT_ROW_HEIGHT', 13.5)
        self.default_column_width = getattr(souren_config, 'EXCEL_DEFAULT_COLUMN_WIDTH', 9)
        self.header_row_height = getattr(souren_config, 'EXCEL_HEADER_ROW_HEIGHT', 20)
        
        # 应用列宽配置（使用安全获取方式）
        self.summary_column_widths = getattr(souren_config, 'EXCEL_SUMMARY_COLUMN_WIDTHS', {})
        self.details_column_widths = getattr(souren_config, 'EXCEL_DETAILS_COLUMN_WIDTHS', {})
        
        # 图表列宽配置（如果不存在则使用默认值）
        if hasattr(souren_config, 'EXCEL_CHART_COLUMN_WIDTHS'):
            self.chart_column_widths = souren_config.EXCEL_CHART_COLUMN_WIDTHS
        else:
            self.chart_column_widths = {'default': 15}
            print(f"⚠️  配置中缺少EXCEL_CHART_COLUMN_WIDTHS，使用默认值")
        
        # 应用颜色配置
        self.border_colors = getattr(souren_config, 'EXCEL_BORDER_COLORS', {"default": "000000"})
        
        # 生成图表颜色的函数（如果存在）
        if hasattr(souren_config, 'generate_chart_colors'):
            self.generate_chart_colors_func = souren_config.generate_chart_colors
        else:
            # 默认颜色生成函数
            self.generate_chart_colors_func = lambda count: plt.cm.tab10(np.linspace(0, 1, min(count, 10)))
        
        # 获取图表配置开关
        self.use_chart_data_extraction = getattr(souren_config, 'USE_CHART_DATA_EXTRACTION', False)
        
        # 根据开关选择配置方式
        if self.use_chart_data_extraction:
            # 使用新的命令模式提取配置
            self.data_extraction_config = getattr(souren_config, 'CHART_DATA_EXTRACTION_CONFIG', {})
            self.chart_steps_filter = []  # 不使用步骤编号筛选
            
            print(f"📊 使用命令模式提取图表数据")
            print(f"   配置内容: {self.data_extraction_config}")
        else:
            # 使用旧的步骤编号筛选配置
            self.data_extraction_config = {}
            self.chart_steps_filter = getattr(souren_config, 'CHART_STEPS_FILTER', [])
            print(f"📊 使用步骤编号筛选配置")
            print(f"   配置内容: {self.chart_steps_filter}")
        
        # 创建样式对象（使用安全获取方式）
        self.header_fill = PatternFill(
            start_color=getattr(souren_config, 'EXCEL_HEADER_FILL_COLOR', "366092"), 
            end_color=getattr(souren_config, 'EXCEL_HEADER_FILL_COLOR', "366092"), 
            fill_type="solid"
        )
        self.header_font = Font(
            color=getattr(souren_config, 'EXCEL_HEADER_FONT_COLOR', "FFFFFF"), 
            bold=True, 
            size=getattr(souren_config, 'EXCEL_HEADER_FONT_SIZE', 11)
        )
        self.query_result_fill = PatternFill(
            start_color=getattr(souren_config, 'EXCEL_QUERY_RESULT_FILL', "E2F0D9"), 
            end_color=getattr(souren_config, 'EXCEL_QUERY_RESULT_FILL', "E2F0D9"), 
            fill_type="solid"
        )
        self.query_result_font = Font(
            color=getattr(souren_config, 'EXCEL_QUERY_RESULT_FONT', "006400"), 
            bold=True
        )
        self.data_font = Font(size=getattr(souren_config, 'EXCEL_DATA_FONT_SIZE', 10))
        
        # 对齐样式
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 图表样式配置
        self.chart_style = {
            'figsize': (10, 6),
            'dpi': 150,
            'grid_alpha': 0.3,
            'font_size_title': 12,
            'font_size_axis': 10,
            'legend_loc': 'best',
            'bar_width': 0.6,
            'bar_alpha': 0.8,
            'colors': plt.cm.Set3(np.linspace(0, 1, 12))  # 使用Set3颜色映射，支持多个循环
        }
    
    def find_latest_json_result(self) -> Optional[str]:
        """查找最新的JSON结果文件"""
        try:
            if not os.path.exists(self.log_dir):
                return None
            
            json_files = []
            for filename in os.listdir(self.log_dir):
                if filename.startswith("souren_results_") and filename.endswith(".json"):
                    filepath = os.path.join(self.log_dir, filename)
                    if os.path.isfile(filepath):
                        mtime = os.path.getmtime(filepath)
                        json_files.append({"path": filepath, "mtime": mtime})
            
            if not json_files:
                return None
            
            json_files.sort(key=lambda x: x["mtime"], reverse=True)
            return json_files[0]["path"]
            
        except Exception:
            return None
    
    def load_json_results(self, json_file: str) -> List[Dict]:
        """加载JSON结果数据"""
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            if not content.strip():
                return []
            
            results = json.loads(content)
            return [results] if isinstance(results, dict) else results
            
        except Exception:
            return []
    
    def convert_to_excel(self, json_file: str = None) -> bool:
        """将JSON结果转换为Excel文件"""
        try:
            if json_file is None:
                json_file = self.find_latest_json_result()
                if json_file is None:
                    print("❌ 未找到JSON结果文件")
                    return False
            
            json_name = os.path.basename(json_file)
            excel_name = json_name.replace('.json', '.xlsx')
            excel_file = os.path.join(os.path.dirname(json_file), excel_name)
            
            # 检查文件是否被锁定，如果是则创建新文件名
            import time
            if os.path.exists(excel_file):
                try:
                    # 尝试重命名文件
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                    backup_file = excel_file.replace('.xlsx', f'_{timestamp}.xlsx')
                    os.rename(excel_file, backup_file)
                    print(f"⚠️ 原文件已存在，已重命名为: {os.path.basename(backup_file)}")
                except:
                    # 如果重命名失败，使用新文件名
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    excel_name = f"souren_results_{timestamp}.xlsx"
                    excel_file = os.path.join(os.path.dirname(json_file), excel_name)
                    print(f"⚠️ 文件被锁定，使用新文件名: {excel_name}")
            
            results = self.load_json_results(json_file)
            if not results:
                print("❌ 未加载到有效的JSON结果数据")
                return False
            
            print(f"📊 开始创建Excel文件: {excel_name}")
            
            # 创建Excel文件
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                self._create_summary_sheet(writer, results)
                self._create_execution_details_sheet(writer, results)
            
            # 应用样式和图表
            self._apply_styles_and_charts(excel_file, results)
            
            print(f"✅ Excel文件已生成: {excel_file}")
            return True
            
        except Exception as e:
            print(f"❌ Excel文件导出失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter, results: List[Dict]):
        """创建汇总表"""
        try:
            valid_results = [r for r in results if r.get('executed_steps', 0) > 0]
            
            if not valid_results:
                df_empty = pd.DataFrame({"状态": ["没有有效的执行记录"]})
                df_empty.to_excel(writer, sheet_name='执行汇总', index=False)
                return
            
            summary_data = []
            for i, result in enumerate(valid_results):
                executed = result.get('executed_steps', 0)
                passed = result.get('passed', 0)
                success_rate = round((passed / executed) * 100, 2) if executed > 0 else 0
                
                summary = {
                    "序号": i + 1,
                    "执行时间": result.get('timestamp_readable', result.get('timestamp', 'N/A')),
                    "SCV文件": result.get('file', 'N/A'),
                    "设备": self._extract_device_name(result),
                    "执行模式": result.get('mode', 'run_all'),
                    "循环次数": result.get('loop_count', 1),
                    "总步骤数": result.get('total_steps', result.get('total_executions', 0)),
                    "已执行步骤": executed,
                    "通过步骤": passed,
                    "失败步骤": result.get('failed', 0),
                    "成功率(%)": success_rate,
                    "总耗时(秒)": round(float(result.get('execution_time', result.get('duration', 0))), 2),
                    "状态": result.get('status', 'N/A'),
                    "状态消息": result.get('message', 'N/A')
                }
                summary_data.append(summary)
            
            if summary_data:
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='执行汇总', index=False)
            
        except Exception as e:
            print(f"创建汇总表失败: {e}")
    
    def _extract_device_name(self, result: Dict) -> str:
        """提取设备名称"""
        device = result.get('device', {})
        if isinstance(device, dict):
            return device.get('name', device.get('address', 'N/A'))
        elif isinstance(device, str):
            return device
        else:
            return 'N/A'
    
    def _create_execution_details_sheet(self, writer: pd.ExcelWriter, results: List[Dict]):
        """创建详细执行记录表"""
        try:
            details_data = []
            
            for result_index, result in enumerate(results):
                if result.get('executed_steps', 0) == 0:
                    continue
                    
                execution_details = result.get('execution_details', [])
                loop_count = result.get('loop_count', 1)
                
                if result.get('mode', 'run_all') != 'loop_info' or loop_count <= 1:
                    for detail in execution_details:
                        details_data.append(self._create_detail_record(detail, result_index, result, 1, 1))
                    continue
                
                # 按循环轮次分组
                loop_groups = {}
                for detail in execution_details:
                    loop_index = detail.get('loop_index', 1)
                    if loop_index not in loop_groups:
                        loop_groups[loop_index] = []
                    loop_groups[loop_index].append(detail)
                
                for loop_index in sorted(loop_groups.keys()):
                    if loop_index > 1:
                        details_data.extend([{}, {}])
                    for detail in loop_groups[loop_index]:
                        details_data.append(self._create_detail_record(
                            detail, result_index, result, loop_index, loop_count
                        ))
            
            if details_data:
                columns = [
                    "汇总序号", "执行时间", "步骤序号", "循环轮次", "总循环数", 
                    "步骤内容", "步骤类型", "执行状态", "执行结果", "耗时(秒)"
                ]
                
                formatted_data = []
                for record in details_data:
                    formatted_record = {}
                    for col in columns:
                        formatted_record[col] = record.get(col, "")
                    formatted_data.append(formatted_record)
                
                df_details = pd.DataFrame(formatted_data, columns=columns)
                df_details.to_excel(writer, sheet_name='详细执行记录', index=False)
            
        except Exception as e:
            print(f"创建详细记录表失败: {e}")
    
    def _create_detail_record(self, detail: Dict, result_index: int, result: Dict, 
                            loop_index: int, total_loops: int) -> Dict:
        """创建单个详细记录"""
        return {
            "汇总序号": result_index + 1,
            "执行时间": result.get('timestamp_readable', result.get('timestamp', 'N/A')),
            "步骤序号": detail.get('original_step', detail.get('step', 'N/A')),
            "循环轮次": loop_index,
            "总循环数": total_loops,
            "步骤内容": detail.get('content', ''),
            "步骤类型": detail.get('type', 'Normal'),
            "执行状态": detail.get('status', 'N/A'),
            "执行结果": str(detail.get('result', 'N/A')),
            "耗时(秒)": round(float(detail.get('duration', 0)), 3)
        }
    
    def _extract_chart_data(self, results: List[Dict]) -> Dict:
        """提取图表数据"""
        if self.use_chart_data_extraction:
            return self._extract_chart_data_by_command(results)
        else:
            return self._extract_chart_data_by_step(results)
    
    def _extract_chart_data_by_command(self, results: List[Dict]) -> Dict:
        """根据配置的命令模式提取图表数据"""
        chart_data = {}
        
        print(f"📊 开始根据命令模式提取数据...")
        print(f"   配置的命令模式: {list(self.data_extraction_config.keys())}")
        
        for result in results:
            execution_details = result.get('execution_details', [])
            loop_count = result.get('loop_count', 1)
            
            for detail in execution_details:
                step_content = detail.get('content', '')
                result_str = str(detail.get('result', ''))
                step_num = detail.get('original_step', detail.get('step', 0))
                loop_index = detail.get('loop_index', 1)
                
                # 检查是否匹配配置的命令模式
                for command_pattern, positions in self.data_extraction_config.items():
                    if command_pattern in step_content:
                        # 提取数字
                        numbers = self._extract_all_numbers(result_str)
                        
                        if numbers:
                            # 为每个提取位置创建数据系列
                            for pos_idx, position in enumerate(positions):
                                if 1 <= position <= len(numbers):
                                    value = numbers[position - 1]
                                    
                                    # 生成数据系列键
                                    series_key = f"{command_pattern}_位置{position}"
                                    
                                    if series_key not in chart_data:
                                        chart_data[series_key] = {
                                            "command": command_pattern,
                                            "position": position,
                                            "data": {},
                                            "loop_indices": set()
                                        }
                                    
                                    # 存储循环数据
                                    loop_key = f"循环{loop_index}"
                                    if loop_key not in chart_data[series_key]["data"]:
                                        chart_data[series_key]["data"][loop_key] = value
                                        chart_data[series_key]["loop_indices"].add(loop_index)
                                    
                                    print(f"    ✅ 提取到数据: {command_pattern} 循环{loop_index} 位置{position} = {value:.3f}")
        
        print(f"📊 共提取了 {len(chart_data)} 个数据系列")
        return chart_data
    
    def _extract_chart_data_by_step(self, results: List[Dict]) -> Dict:
        """按步骤编号提取图表数据"""
        chart_data = {}
        
        print(f"📊 使用步骤编号筛选数据...")
        print(f"   筛选步骤: {self.chart_steps_filter}")
        
        for result in results:
            execution_details = result.get('execution_details', [])
            
            for detail in execution_details:
                step_content = detail.get('content', '')
                result_str = str(detail.get('result', ''))
                step_num = detail.get('original_step', detail.get('step', 0))
                loop_index = detail.get('loop_index', 1)
                
                # 根据配置筛选步骤
                if step_num not in self.chart_steps_filter:
                    continue
                
                # 只处理查询命令且有结果的数据
                if '?' in step_content and result_str and result_str != 'N/A':
                    numbers = self._extract_all_numbers(result_str)
                    
                    if numbers:
                        step_key = f"步骤{step_num}"
                        loop_key = f"循环{loop_index}"
                        
                        if step_key not in chart_data:
                            chart_data[step_key] = {
                                "step_num": int(step_num),
                                "content": step_content.replace('?', '').strip(),
                                "loops": {}
                            }
                        
                        # 存储该循环的数据点
                        if loop_key not in chart_data[step_key]["loops"]:
                            chart_data[step_key]["loops"][loop_key] = {
                                "loop_index": loop_index,
                                "data_points": numbers
                            }
        
        print(f"📊 根据步骤编号筛选了 {len(chart_data)} 个步骤的数据")
        return chart_data
    
    def _extract_all_numbers(self, text: str) -> List[float]:
        """从文本中提取所有数字，保持原值"""
        try:
            if not text:
                return []
            
            clean_text = text.strip()
            
            # 检查是否包含错误信息
            error_keywords = ["仪器通信错误", "VI_ERROR_TMO", "Timeout", "通信失败", "错误", "ERROR"]
            if any(keyword in clean_text for keyword in error_keywords):
                return []
            
            # 检查是否为空或无效响应
            if clean_text in ['', 'N/A', 'None', 'null', '[]', '{}']:
                return []
            
            # 处理逗号分隔的数组格式（这是最常见的情况）
            if ',' in clean_text:
                parts = [part.strip() for part in clean_text.split(',')]
                numbers = []
                for part in parts:
                    try:
                        # 尝试直接转换为数字
                        num = float(part)
                        numbers.append(num)
                    except ValueError:
                        # 如果不是纯数字，尝试提取数字
                        pattern = r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?'
                        matches = re.findall(pattern, part)
                        for match in matches:
                            try:
                                numbers.append(float(match))
                            except ValueError:
                                continue
                if numbers:
                    return numbers
            
            # 尝试直接解析为单个数字
            try:
                num = float(clean_text)
                return [num]
            except ValueError:
                pass
            
            # 常规提取所有数字
            pattern = r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?'
            matches = re.findall(pattern, clean_text)
            
            numbers = []
            for match in matches:
                try:
                    number = float(match)
                    numbers.append(number)
                except ValueError:
                    continue
            
            return numbers
        except Exception as e:
            print(f"❌ 提取数字失败: {e}, 文本: {text[:100]}...")
            return []
    
    def _create_bar_chart(self, command: str, position: int, 
                         loop_data: Dict[str, float], loop_indices: set) -> BytesIO:
        """
        创建柱状图显示不同循环的数据
        """
        try:
            # 准备数据
            sorted_loop_keys = sorted(loop_data.keys(), key=lambda x: int(x.replace('循环', '')))
            loop_values = []
            loop_labels = []
            
            for loop_key in sorted_loop_keys:
                loop_labels.append(loop_key)
                loop_values.append(loop_data[loop_key])
            
            if not loop_values:
                return None
            
            # 创建图表
            fig, ax = plt.subplots(figsize=self.chart_style['figsize'])
            
            # 为每个循环使用不同的颜色
            colors = self.chart_style['colors'][:len(loop_values)]
            
            # 创建柱状图
            bars = ax.bar(range(len(loop_values)), loop_values, 
                         color=colors, 
                         alpha=self.chart_style['bar_alpha'],
                         width=self.chart_style['bar_width'])
            
            # 添加数值标签
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{height:.3f}',
                       ha='center', va='bottom',
                       fontsize=10)
            
            # 设置图表属性
            ax.set_xlabel('循环轮次', fontsize=self.chart_style['font_size_axis'])
            ax.set_ylabel('数值', fontsize=self.chart_style['font_size_axis'])
            
            # 提取命令名称
            command_name = self._get_command_name(command)
            title = f'{command_name} (位置{position})'
            ax.set_title(title, fontsize=self.chart_style['font_size_title'])
            
            # 设置x轴刻度
            ax.set_xticks(range(len(loop_labels)))
            ax.set_xticklabels(loop_labels)
            
            # 添加网格
            ax.grid(True, alpha=self.chart_style['grid_alpha'], axis='y')
            
            # 调整布局
            plt.tight_layout()
            
            # 将图表保存到BytesIO
            img_data = BytesIO()
            plt.savefig(img_data, format='png', dpi=self.chart_style['dpi'], 
                       bbox_inches='tight', facecolor='white')
            img_data.seek(0)
            plt.close(fig)  # 关闭图形释放内存
            
            return img_data
            
        except Exception as e:
            print(f"❌ 创建柱状图失败: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _get_command_name(self, command: str) -> str:
        """从命令中提取简化的名称"""
        # 移除查询符号
        clean_command = command.replace('?', '').strip()
        
        # 提取关键部分
        if 'BLER:UL' in clean_command or 'UL:RESult' in clean_command:
            return 'UL BLER'
        elif 'BLER:DL' in clean_command or 'DL:RESult' in clean_command:
            return 'DL BLER'
        elif 'TXP:AVG' in clean_command:
            return 'TXP AVG'
        elif 'TXP:MIN' in clean_command:
            return 'TXP MIN'
        elif 'TXP:MAX' in clean_command:
            return 'TXP MAX'
        
        # 提取最后一部分作为名称
        parts = clean_command.split(':')
        if parts:
            return parts[-1]
        
        return clean_command[:20]
    
    def _apply_styles_and_charts(self, excel_file: str, results: List[Dict]):
        """应用样式和图表到Excel文件"""
        try:
            wb = load_workbook(excel_file)
            
            # 1. 格式化执行汇总表
            if '执行汇总' in wb.sheetnames:
                self._format_summary_sheet(wb['执行汇总'])
            
            # 2. 格式化详细执行记录表
            if '详细执行记录' in wb.sheetnames:
                self._format_details_sheet(wb['详细执行记录'])
            
            # 3. 提取图表数据
            chart_data = self._extract_chart_data(results)
            
            if chart_data:
                # 4. 创建数据分析图表工作表
                self._create_chart_sheet(wb, chart_data)
            else:
                # 创建空的数据分析图表工作表
                if '数据分析图表' in wb.sheetnames:
                    wb.remove(wb['数据分析图表'])
                chart_ws = wb.create_sheet(title='数据分析图表')
                chart_ws['A1'] = "数据分析图表"
                
                if self.use_chart_data_extraction:
                    chart_ws['A2'] = "⚠️ 没有找到配置命令的查询数据"
                    chart_ws['A3'] = f"配置的命令模式: {list(self.data_extraction_config.keys())}"
                    print(f"⚠️ 没有找到配置命令的查询数据")
                    print(f"   配置的命令模式: {list(self.data_extraction_config.keys())}")
                else:
                    chart_ws['A2'] = "⚠️ 没有找到配置步骤的查询数据"
                    chart_ws['A3'] = f"配置的步骤编号: {self.chart_steps_filter}"
                    print(f"⚠️ 没有找到配置步骤的查询数据")
            
            # 5. 调整列宽行高
            self._adjust_dimensions(wb)
            
            wb.save(excel_file)
            
        except Exception as e:
            print(f"❌ 应用样式和图表失败: {e}")
            import traceback
            traceback.print_exc()
    
    def _create_chart_sheet(self, wb, chart_data):
        """创建图表工作表"""
        try:
            # 移除旧的图表工作表
            if '数据分析图表' in wb.sheetnames:
                wb.remove(wb['数据分析图表'])
            
            # 创建新的数据分析图表工作表
            chart_ws = wb.create_sheet(title='数据分析图表')
            
            # 1. 添加标题
            title = "数据分析图表（基于命令模式提取）"
            config_info = f"数据提取配置: {self.data_extraction_config}"
            
            chart_ws['A1'] = title
            chart_ws['A1'].font = Font(size=14, bold=True, color=self.header_fill.start_color.rgb)
            chart_ws.merge_cells('A1:L1')
            
            # 显示配置信息
            chart_ws['A2'] = config_info
            chart_ws['A2'].font = Font(size=10, italic=True)
            
            print(f"📊 开始创建柱状图...")
            
            # 2. 为每个数据系列创建柱状图
            current_row = 5
            charts_created = 0
            
            for series_key, series_info in chart_data.items():
                command = series_info["command"]
                position = series_info["position"]
                loop_data = series_info["data"]
                
                if not loop_data:
                    continue
                
                print(f"  📊 处理命令: {command}, 位置: {position}")
                print(f"    循环数据: {loop_data}")
                
                # 创建柱状图
                img_data = self._create_bar_chart(command, position, loop_data, series_info["loop_indices"])
                
                if img_data:
                    # 创建Excel图像对象
                    img = ExcelImage(img_data)
                    
                    # 调整图像大小
                    img.width = 1200  # 像素宽度
                    img.height = 600  # 像素高度
                    
                    # 将图像插入到Excel中
                    chart_cell = f"A{current_row}"
                    chart_ws.add_image(img, chart_cell)
                    
                    # 在图表下方添加简要说明
                    explanation_row = current_row + 20
                    chart_ws.cell(row=explanation_row, column=1, value=f"命令: {command}")
                    chart_ws.cell(row=explanation_row + 1, column=1, value=f"提取位置: {position}")
                    chart_ws.cell(row=explanation_row + 2, column=1, value=f"循环数: {len(loop_data)}")
                    
                    # 显示具体数值
                    chart_ws.cell(row=explanation_row + 3, column=1, value="具体数值:")
                    value_row = explanation_row + 4
                    for loop_key, value in sorted(loop_data.items()):
                        chart_ws.cell(row=value_row, column=1, value=f"{loop_key}: {value:.3f}")
                        value_row += 1
                    
                    # 更新下一张图表的起始位置
                    current_row = value_row + 3
                    charts_created += 1
                else:
                    print(f"  ❌ 创建图表失败: {command}")
            
            if charts_created == 0:
                chart_ws['A5'] = "⚠️ 没有足够的数据来创建图表"
                chart_ws['A6'] = "可能原因:"
                chart_ws['A7'] = "1. 数据点数量不足"
                chart_ws['A8'] = "2. 查询命令没有返回有效的数字数据"
                chart_ws['A9'] = "3. 配置的命令模式不匹配: {list(self.data_extraction_config.keys())}"
                
                print("⚠️ 没有足够的数据来创建图表")
            else:
                print(f"✅ 成功创建了 {charts_created} 个柱状图并插入到Excel")
            
            # 3. 设置列宽
            for col_letter, width in self.chart_column_widths.items():
                if col_letter == 'default':
                    # 设置默认列宽
                    for col in range(1, chart_ws.max_column + 1):
                        col_letter = get_column_letter(col)
                        if col_letter not in self.chart_column_widths:
                            chart_ws.column_dimensions[col_letter].width = width
                else:
                    if col_letter in chart_ws.column_dimensions:
                        chart_ws.column_dimensions[col_letter].width = width
            
        except Exception as e:
            print(f"❌ 创建图表工作表失败: {e}")
            import traceback
            traceback.print_exc()
    
    def _format_summary_sheet(self, ws):
        """格式化汇总表"""
        try:
            # 格式化表头
            for cell in ws[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.border = self._create_border(self.border_colors.get("default", "000000"))
            
            # 格式化数据行
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = self._create_border(self.border_colors.get("default", "000000"))
                    cell.font = self.data_font
                    
                    # 设置对齐方式
                    col_idx = cell.column
                    if col_idx in [1, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
                        cell.alignment = self.center_align
                    elif col_idx == 2:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = self.left_align
                
                # 设置行高
                ws.row_dimensions[row[0].row].height = self.default_row_height
            
        except Exception as e:
            print(f"❌ 格式化汇总表失败: {e}")
    
    def _format_details_sheet(self, ws):
        """格式化详细执行记录表"""
        try:
            # 格式化表头
            for cell in ws[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.border = self._create_border(self.border_colors.get("default", "000000"))
            
            # 设置表头行高
            ws.row_dimensions[1].height = self.header_row_height
            
            # 处理数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                if not any(cell.value for cell in row):
                    continue
                
                # 检查是否是查询命令
                step_content_cell = row[5]  # F列是步骤内容
                is_query = step_content_cell.value and '?' in str(step_content_cell.value)
                
                for cell in row:
                    cell.border = self._create_border(self.border_colors.get("default", "000000"))
                    cell.font = self.data_font
                    
                    # 设置对齐方式
                    col_idx = cell.column
                    if col_idx in [1, 3, 4, 5, 7, 8, 10]:
                        cell.alignment = self.center_align
                    elif col_idx == 2:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = self.left_align
                
                # 高亮显示查询结果行
                if is_query:
                    for cell in row:
                        cell.fill = self.query_result_fill
                        cell.font = self.query_result_font
                
                # 设置数据行高
                ws.row_dimensions[row_idx].height = self.default_row_height
            
        except Exception as e:
            print(f"❌ 格式化详细记录表失败: {e}")
    
    def _create_border(self, color: str) -> Border:
        """创建边框"""
        side = Side(style='thin', color=color)
        return Border(left=side, right=side, top=side, bottom=side)
    
    def _adjust_dimensions(self, wb):
        """调整列宽和行高"""
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 设置行高
                for row in range(1, ws.max_row + 1):
                    if row == 1:
                        # 表头行高
                        ws.row_dimensions[row].height = self.header_row_height
                    else:
                        # 数据行高
                        ws.row_dimensions[row].height = self.default_row_height
                
                # 根据工作表类型应用不同的列宽设置
                if sheet_name == '执行汇总':
                    # 首先设置默认列宽
                    for col in ws.columns:
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = self.default_column_width
                    
                    # 然后应用特殊列宽配置
                    for col_letter, width in self.summary_column_widths.items():
                        if col_letter in ws.column_dimensions:
                            ws.column_dimensions[col_letter].width = width
                            
                elif sheet_name == '详细执行记录':
                    # 首先设置默认列宽
                    for col in ws.columns:
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = self.default_column_width
                    
                    # 然后应用特殊列宽配置
                    for col_letter, width in self.details_column_widths.items():
                        if col_letter in ws.column_dimensions:
                            ws.column_dimensions[col_letter].width = width
                            
                elif sheet_name == '数据分析图表':
                    # 应用图表列宽配置
                    for col_letter, width in self.chart_column_widths.items():
                        if col_letter == 'default':
                            # 设置默认列宽
                            for col in range(1, ws.max_column + 1):
                                col_letter = get_column_letter(col)
                                if col_letter not in self.chart_column_widths:
                                    ws.column_dimensions[col_letter].width = width
                        elif col_letter in ws.column_dimensions:
                            ws.column_dimensions[col_letter].width = width
                            
                else:
                    # 其他工作表设置默认列宽
                    for col in ws.columns:
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = self.default_column_width
                
                # 启用筛选（除了图表工作表）
                if sheet_name != '数据分析图表' and ws.max_row > 1:
                    try:
                        ws.auto_filter.ref = ws.dimensions
                    except:
                        pass
            
            print(f"✅ 已应用所有Excel导出配置")
            print(f"   - 当前模式: {'命令模式提取' if self.use_chart_data_extraction else '步骤编号筛选'}")
            print(f"   - 图表类型: 柱状图（每个循环不同颜色）")
            
        except Exception as e:
            print(f"❌ 调整列宽行高失败: {e}")


# ==============================================
# ↓↓↓ 汇总功能（优化版） - 添加横向对比图表
# ==============================================
def create_summary_excel(main_execution_dir: str, all_results: List[Dict]) -> Optional[str]:
    """
    创建汇总Excel文件 - 优化版：包含横向对比图表
    展示多个配置的循环数据对比，一眼看出哪个配置出问题
    """
    try:
        print(f"\n{'='*60}")
        print("📊 开始创建图表汇总Excel文件...")
        print(f"{'='*60}")
        
        if not main_execution_dir or not os.path.exists(main_execution_dir):
            print("❌ 主执行目录不存在，无法创建汇总文件")
            return None
        
        if not all_results or len(all_results) == 0:
            print("❌ 没有可汇总的结果数据")
            return None
        
        # 从souren_config导入配置
        try:
            from souren_config import (
                USE_CHART_DATA_EXTRACTION,
                CHART_DATA_EXTRACTION_CONFIG,
                CHART_STEPS_FILTER,
                PYTHON_SCRIPT_NAME,
                LOOP_COUNT
            )
        except ImportError as e:
            print(f"❌ 导入配置失败: {e}")
            return None
        
        print(f"📊 配置信息:")
        print(f"   USE_CHART_DATA_EXTRACTION: {USE_CHART_DATA_EXTRACTION}")
        print(f"   PYTHON_SCRIPT_NAME 数量: {len(PYTHON_SCRIPT_NAME)}")
        print(f"   LOOP_COUNT: {LOOP_COUNT}")
        
        if USE_CHART_DATA_EXTRACTION:
            print(f"   CHART_DATA_EXTRACTION_CONFIG: {CHART_DATA_EXTRACTION_CONFIG}")
        else:
            print(f"   CHART_STEPS_FILTER: {CHART_STEPS_FILTER}")
        
        # 收集所有配置的信息
        config_info_map = {}
        max_loop_count = 0  # 记录最大循环次数
        
        for result in all_results:
            script_name = result.get('script_name', '未知脚本')
            parameters = result.get('parameters', {})
            script_dir = result.get('script_directory', '')
            
            # 构建配置标识
            config_id = f"{script_name}"
            if parameters:
                band = parameters.get('band', '')
                bw = parameters.get('bw', '')
                scs = parameters.get('scs', '')
                range_val = parameters.get('range', '')
                line_loss = parameters.get('lineLoss', '')
                # 创建简短的配置名称
                short_id = f"{script_name.replace('.py', '')}"
                if line_loss:
                    # 处理lineLoss值，提取整数部分
                    try:
                        if isinstance(line_loss, str):
                            line_loss_int = int(float(line_loss))
                        else:
                            line_loss_int = int(line_loss)
                        short_id += f"_{line_loss_int}"
                    except:
                        short_id += f"_{line_loss}"
                if band:
                    short_id += f"_b{band}"
                if bw:
                    short_id += f"_bw{bw}"
                if scs:
                    short_id += f"_scs{scs}"
                if range_val:
                    short_id += f"_{range_val}"
                config_id = short_id
            
            # 获取循环次数
            loop_count = result.get('loop_count', 1)
            if loop_count > max_loop_count:
                max_loop_count = loop_count
            
            # 保存配置信息
            config_info_map[config_id] = {
                'script_name': script_name,
                'parameters': parameters,
                'script_dir': script_dir,
                'loop_count': loop_count,
                'result_data': result  # 直接保存结果数据
            }
        
        print(f"📊 找到 {len(config_info_map)} 个配置信息")
        print(f"📊 最大循环次数: {max_loop_count}")
        
        # 提取所有配置的图表数据
        all_chart_data = {}
        data_extraction_errors = []
        
        for config_id, config_info in config_info_map.items():
            script_dir = config_info.get('script_dir', '')
            result_data = config_info.get('result_data', {})
            
            print(f"🔍 处理配置: {config_id}")
            
            # 方法1：首先尝试从result_data中提取
            if result_data:
                # 提取图表数据
                chart_data = {}
                
                if USE_CHART_DATA_EXTRACTION:
                    # 使用命令模式提取
                    chart_data = _extract_chart_data_by_command_custom(
                        [result_data], 
                        CHART_DATA_EXTRACTION_CONFIG
                    )
                else:
                    # 使用步骤编号筛选
                    chart_data = _extract_chart_data_by_step_custom(
                        [result_data],
                        CHART_STEPS_FILTER
                    )
                
                if chart_data:
                    all_chart_data[config_id] = {
                        'config_info': config_info,
                        'chart_data': chart_data,
                        'source': 'result_data'
                    }
                    print(f"✅ 从结果数据中提取到图表数据: {len(chart_data)} 个数据项")
                    continue
            
            # 方法2：如果result_data中没有图表数据，尝试从JSON文件中提取
            if not os.path.exists(script_dir):
                print(f"❌ 脚本目录不存在: {script_dir}")
                data_extraction_errors.append(f"目录不存在: {config_id}")
                continue
            
            # 查找该目录下的所有JSON文件
            json_files = []
            for file in os.listdir(script_dir):
                if file.endswith(".json") and not file.startswith("summary"):
                    json_file = os.path.join(script_dir, file)
                    json_files.append(json_file)
            
            if not json_files:
                print(f"❌ 在目录 {script_dir} 中未找到JSON结果文件")
                data_extraction_errors.append(f"无JSON文件: {config_id}")
                continue
            
            # 使用最新的JSON文件
            json_file = json_files[-1]  # 通常最后一个是最新的
            
            try:
                # 读取JSON文件
                with open(json_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                if not content.strip():
                    print(f"❌ JSON文件为空: {json_file}")
                    data_extraction_errors.append(f"JSON文件为空: {config_id}")
                    continue
                
                # 解析JSON
                results_data = json.loads(content)
                if isinstance(results_data, dict):
                    results_data = [results_data]
                
                print(f"✅ 成功加载 {config_id} 的JSON数据")
                
                # 提取图表数据
                chart_data = {}
                
                if USE_CHART_DATA_EXTRACTION:
                    # 使用命令模式提取
                    chart_data = _extract_chart_data_by_command_custom(
                        results_data, 
                        CHART_DATA_EXTRACTION_CONFIG
                    )
                else:
                    # 使用步骤编号筛选
                    chart_data = _extract_chart_data_by_step_custom(
                        results_data,
                        CHART_STEPS_FILTER
                    )
                
                if chart_data:
                    all_chart_data[config_id] = {
                        'config_info': config_info,
                        'chart_data': chart_data,
                        'json_file': json_file,
                        'source': 'json_file'
                    }
                    print(f"✅ 从JSON文件中提取到图表数据: {len(chart_data)} 个数据项")
                else:
                    print(f"⚠️  未提取到 {config_id} 的图表数据")
                    data_extraction_errors.append(f"无图表数据: {config_id}")
                    
            except Exception as e:
                error_msg = f"处理 {config_id} 失败: {e}"
                print(f"❌ {error_msg}")
                data_extraction_errors.append(error_msg)
                continue
        
        if not all_chart_data:
            print("❌ 没有提取到任何图表数据")
            
            if data_extraction_errors:
                print("📊 数据提取错误详情:")
                for error in data_extraction_errors:
                    print(f"  - {error}")
            
            return None
        
        print(f"📊 成功提取到 {len(all_chart_data)} 个配置的图表数据")
        
        # 创建汇总Excel文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        summary_excel_name = f"summary_charts_{timestamp}.xlsx"
        summary_excel_path = os.path.join(main_execution_dir, summary_excel_name)
        
        # 使用openpyxl创建Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as ExcelImage
        
        wb = Workbook()
        
        # 1. 创建图表汇总表
        ws = wb.active
        ws.title = "图表汇总"
        
        # 设置标题
        ws['A1'] = "测试数据图表汇总"
        ws['A1'].font = Font(size=16, bold=True, color="366092")
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        ws['A3'] = f"配置数量: {len(all_chart_data)}"
        ws['B3'] = f"最大循环数: {max_loop_count}"
        
        current_row = 5
        
        # 2. 为每个配置创建单独的图表
        charts_created = 0
        
        for config_id, data_info in all_chart_data.items():
            config_info = data_info['config_info']
            chart_data = data_info['chart_data']
            script_name = config_info['script_name']
            parameters = config_info['parameters']
            
            # 添加配置标题
            ws.cell(row=current_row, column=1, value=f"配置: {config_id}")
            ws.cell(row=current_row, column=1).font = Font(size=12, bold=True)
            current_row += 1
            
            # 添加参数信息
            if parameters:
                param_info = ""
                if 'band' in parameters:
                    param_info += f"band:{parameters['band']} "
                if 'bw' in parameters:
                    param_info += f"bw:{parameters['bw']} "
                if 'scs' in parameters:
                    param_info += f"scs:{parameters['scs']} "
                if 'range' in parameters:
                    param_info += f"range:{parameters['range']}"
                
                ws.cell(row=current_row, column=1, value=f"参数: {param_info}")
                current_row += 1
            
            # 处理每个数据项
            for data_key, data_item in chart_data.items():
                if 'data' in data_item:  # 命令模式
                    loop_data = data_item['data']
                    command = data_item['command']
                    position = data_item['position']
                    
                    # 创建图表
                    img_data = _create_chart_image(
                        loop_data, 
                        command, 
                        position,
                        config_id
                    )
                    
                    if img_data:
                        # 创建Excel图像对象
                        img = ExcelImage(img_data)
                        img.width = 600  # 像素宽度
                        img.height = 300  # 像素高度
                        
                        # 将图像插入到Excel中
                        chart_cell = f"A{current_row}"
                        ws.add_image(img, chart_cell)
                        
                        # 添加图表说明
                        ws.cell(row=current_row + 15, column=1, value=f"命令: {command}")
                        ws.cell(row=current_row + 16, column=1, value=f"提取位置: {position}")
                        
                        # 显示循环数据
                        value_row = current_row + 17
                        ws.cell(row=value_row, column=1, value="循环数据:")
                        for loop_key, value in sorted(loop_data.items()):
                            ws.cell(row=value_row, column=2, value=f"{loop_key}: {value:.3f}")
                            value_row += 1
                        
                        current_row = value_row + 2
                        charts_created += 1
                
                elif 'loops' in data_item:  # 步骤模式
                    loops_data = data_item['loops']
                    step_num = data_item.get('step_num', '?')
                    step_content = data_item.get('content', '')
                    
                    # 转换为loop_data格式
                    loop_data = {}
                    for loop_key, loop_info in loops_data.items():
                        data_points = loop_info.get('data_points', [])
                        if data_points:
                            loop_data[loop_key] = data_points[0]
                    
                    if loop_data:
                        # 创建图表
                        img_data = _create_chart_image(
                            loop_data, 
                            step_content, 
                            step_num,
                            config_id
                        )
                        
                        if img_data:
                            # 创建Excel图像对象
                            img = ExcelImage(img_data)
                            img.width = 600
                            img.height = 300
                            
                            # 将图像插入到Excel中
                            chart_cell = f"A{current_row}"
                            ws.add_image(img, chart_cell)
                            
                            # 添加图表说明
                            ws.cell(row=current_row + 15, column=1, value=f"步骤: {step_num}")
                            ws.cell(row=current_row + 16, column=1, value=f"内容: {step_content}")
                            
                            # 显示循环数据
                            value_row = current_row + 17
                            ws.cell(row=value_row, column=1, value="循环数据:")
                            for loop_key, value in sorted(loop_data.items()):
                                ws.cell(row=value_row, column=2, value=f"{loop_key}: {value:.3f}")
                                value_row += 1
                            
                            current_row = value_row + 2
                            charts_created += 1
            
            # 配置之间添加分隔行
            current_row += 2
        
        # 3. 创建横向对比图表（优化显示）
        print(f"\n📊 开始创建横向对比图表...")
        
        # 创建新的横向对比工作表
        comparison_ws = wb.create_sheet(title="横向对比")
        
        # 设置标题
        comparison_ws['A1'] = "配置横向对比图表"
        comparison_ws['A1'].font = Font(size=16, bold=True, color="366092")
        comparison_ws.merge_cells('A1:H1')
        
        comparison_ws['A2'] = "将所有配置的相同指标放在一起对比，一眼看出哪个配置出问题"
        comparison_ws['A2'].font = Font(size=10, italic=True)
        
        comparison_current_row = 5
        
        # 收集所有配置的相同命令的数据
        if USE_CHART_DATA_EXTRACTION and CHART_DATA_EXTRACTION_CONFIG:
            # 对于配置的每个命令和位置，收集所有配置的数据
            for command_pattern, positions in CHART_DATA_EXTRACTION_CONFIG.items():
                for position in positions:
                    # 收集所有配置的该命令位置数据
                    comparison_data = {}
                    
                    for config_id, data_info in all_chart_data.items():
                        chart_data = data_info['chart_data']
                        config_info = data_info['config_info']
                        
                        # 查找该命令位置的数据
                        series_key = f"{command_pattern}_位置{position}"
                        if series_key in chart_data:
                            loop_data = chart_data[series_key]["data"]
                            comparison_data[config_id] = {
                                'config_info': config_info,
                                'loop_data': loop_data
                            }
                    
                    # 如果有多个配置的数据，创建横向对比图表
                    if len(comparison_data) > 1:
                        print(f"  📊 创建横向对比图表: {command_pattern} 位置{position}")
                        print(f"     涉及配置: {list(comparison_data.keys())}")
                        
                        # 创建横向对比图表图像
                        comparison_img_data = _create_comparison_chart_image(
                            comparison_data,
                            command_pattern,
                            position,
                            len(PYTHON_SCRIPT_NAME),  # 传递配置数量
                            LOOP_COUNT  # 传递循环次数
                        )
                        
                        if comparison_img_data:
                            # 创建Excel图像对象
                            img = ExcelImage(comparison_img_data)
                            
                            # 根据配置数量动态调整图像大小
                            config_count = len(comparison_data)
                            img_width = min(2000, 1000 + config_count * 60)  # 动态宽度
                            img_height = min(1200, 600 + config_count * 30)  # 动态高度
                            img.width = img_width
                            img.height = img_height
                            
                            # 将图像插入到Excel中
                            chart_cell = f"A{comparison_current_row}"
                            comparison_ws.add_image(img, chart_cell)
                            
                            # 在图表下方添加简化的命令信息
                            info_row = comparison_current_row + int(img_height / 15) + 5  # 根据图表高度调整
                            comparison_ws.cell(row=info_row, column=1, 
                                            value=f"命令: {_simplify_command_name(command_pattern)}")
                            comparison_ws.cell(row=info_row + 1, column=1, 
                                            value=f"提取位置: {position}")
                            comparison_ws.cell(row=info_row + 2, column=1, 
                                            value=f"配置数量: {len(comparison_data)}")
                            
                            # 更新下一张图表的起始位置
                            comparison_current_row = info_row + 10
        
        # 设置列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        comparison_ws.column_dimensions['A'].width = 30
        comparison_ws.column_dimensions['B'].width = 20
        comparison_ws.column_dimensions['C'].width = 15
        
        # 4. 保存Excel文件
        wb.save(summary_excel_path)
        
        print(f"\n{'='*60}")
        print(f"✅ 图表汇总Excel文件创建成功!")
        print(f"📁 文件路径: {summary_excel_path}")
        print(f"📊 图表汇总: {charts_created} 个图表")
        print(f"📈 横向对比: 已创建")
        print(f"{'='*60}")
        
        return summary_excel_path
        
    except Exception as e:
        print(f"❌ 创建图表汇总Excel文件失败: {e}")
        import traceback
        traceback.print_exc()
        return None


def _create_comparison_chart_image(comparison_data: Dict, command: str, position: int, 
                                  config_count: int, loop_count: int) -> Optional[BytesIO]:
    """
    创建横向对比图表图像，显示多个配置的循环数据对比
    优化：完全移除配置标题框，动态调整图表大小以适应更多数据
    """
    try:
        # 准备数据
        config_ids = []
        loop_labels = []
        
        # 确定所有配置共有的循环标签
        all_loops = set()
        for config_id, data in comparison_data.items():
            loop_data = data['loop_data']
            all_loops.update(loop_data.keys())
        
        loop_labels = sorted(list(all_loops), key=lambda x: int(x.replace('循环', '')))
        
        # 为每个循环创建数据系列
        loop_series = {}
        for loop_label in loop_labels:
            loop_series[loop_label] = []
        
        # 收集每个配置的循环数据
        for config_id, data in comparison_data.items():
            config_ids.append(config_id)
            loop_data = data['loop_data']
            
            for loop_label in loop_labels:
                value = loop_data.get(loop_label, 0)  # 如果没有数据，默认为0
                loop_series[loop_label].append(value)
        
        if not config_ids or not loop_labels:
            return None
        
        # 动态计算图表大小
        # 基础尺寸 + 根据配置数量和循环次数调整
        base_width = 14
        base_height = 9
        width_multiplier = 0.6
        height_multiplier = 0.4
        
        # 计算动态尺寸
        dynamic_width = base_width + (config_count * width_multiplier)
        dynamic_height = base_height + (loop_count * height_multiplier)
        
        # 设置最大和最小尺寸
        max_width = 30
        max_height = 20
        min_width = 14
        min_height = 9
        
        fig_width = min(max_width, max(min_width, dynamic_width))
        fig_height = min(max_height, max(min_height, dynamic_height))
        
        print(f"  📐 图表尺寸: {fig_width:.1f} x {fig_height:.1f} (配置: {config_count}, 循环: {loop_count})")
        
        # 创建图表 - 使用动态尺寸，设置更高的DPI以获得更清晰的图像
        fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=180)
        
        # 为每个循环创建柱状图
        x = np.arange(len(config_ids))
        
        # 动态调整柱状图宽度，增加间隔
        bar_width = 0.9 / max(len(loop_labels), 1)  # 防止除零
        bar_width = min(0.35, bar_width)  # 设置最大宽度
        
        # 增加配置之间的间隔
        group_spacing = 1.3
        x_positions = x * group_spacing
        
        # 为每个循环使用不同颜色
        colors = plt.cm.Set3(np.linspace(0, 1, len(loop_labels)))
        
        # 存储每个配置的所有柱子的位置
        config_bars = {}
        
        for i, (loop_label, color) in enumerate(zip(loop_labels, colors)):
            values = loop_series[loop_label]
            
            # 计算每个柱子的位置，增加间隔
            bar_positions = x_positions + i * bar_width - (len(loop_labels)-1) * bar_width / 2
            bars = ax.bar(bar_positions, values, bar_width, label=loop_label, color=color, alpha=0.85)
            
            # 存储每个配置的柱子位置
            for j, bar in enumerate(bars):
                config_id = config_ids[j]
                if config_id not in config_bars:
                    config_bars[config_id] = []
                config_bars[config_id].append(bar)
            
            # 添加数值标签，显示3位小数
            for bar in bars:
                height = bar.get_height()
                if height > 0:  # 只显示非零值
                    # 根据数值大小调整字体大小
                    label_fontsize = 10 if config_count <= 15 else 8
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           f'{height:.3f}',  # 显示3位小数
                           ha='center', va='bottom',
                           fontsize=label_fontsize,
                           fontweight='bold',
                           color='black',
                           bbox=dict(boxstyle="round,pad=0.1", facecolor="white", edgecolor="none", alpha=0.7))
        
        # 移除配置标题框 - 不再为每个配置添加配置标签和矩形框
        # 之前这里有一段代码为每个配置添加矩形框和配置标签，现在完全移除
        
        # 设置图表属性 - 完全移除任何标题
        ax.set_xlabel('配置', fontsize=13, fontweight='bold')
        ax.set_ylabel('数值', fontsize=13, fontweight='bold')
        
        # 设置x轴刻度
        ax.set_xticks(x_positions)
        
        # 根据配置数量调整x轴标签的旋转角度和字体大小
        if config_count > 15:
            rotation_angle = 60
            font_size = 9
        elif config_count > 8:
            rotation_angle = 45
            font_size = 10
        else:
            rotation_angle = 0
            font_size = 11
            
        ax.set_xticklabels(config_ids, rotation=rotation_angle, 
                          ha='right' if rotation_angle > 0 else 'center', 
                          fontsize=font_size, fontweight='bold')
        
        # 设置y轴刻度
        ax.tick_params(axis='y', labelsize=11)
        
        # 添加网格
        ax.grid(True, alpha=0.4, axis='y', linestyle='--', linewidth=0.5)
        
        # 添加图例
        if len(loop_labels) > 0:
            legend_fontsize = 10 if loop_count <= 10 else 9
            legend_title_fontsize = 11 if loop_count <= 10 else 10
            
            ax.legend(title='循环轮次', loc='upper left', bbox_to_anchor=(1.02, 1), 
                     fontsize=legend_fontsize, title_fontsize=legend_title_fontsize, 
                     framealpha=0.9, frameon=True, edgecolor='gray')
        
        # 动态调整布局 - 确保图表足够大
        right_margin = 0.78 if loop_count <= 10 else 0.82
        bottom_margin = 0.18 if config_count <= 10 else 0.25
        top_margin = 0.95  # 增加顶部边距，确保没有标题
        
        plt.subplots_adjust(right=right_margin, top=top_margin, bottom=bottom_margin, left=0.12)
        
        # 将图表保存到BytesIO
        img_data = BytesIO()
        plt.savefig(img_data, format='png', dpi=180, 
                   bbox_inches='tight', facecolor='white', pad_inches=0.3)
        img_data.seek(0)
        plt.close(fig)  # 关闭图形释放内存
        
        return img_data
        
    except Exception as e:
        print(f"❌ 创建横向对比图表失败: {e}")
        import traceback
        traceback.print_exc()
        return None


def _simplify_command_name(command: str) -> str:
    """简化命令名称"""
    if 'BLER:UL' in command or 'UL:RESult' in command:
        return 'UL BLER'
    elif 'BLER:DL' in command or 'DL:RESult' in command:
        return 'DL BLER'
    elif 'TXP:AVG' in command:
        return 'TXP AVG'
    elif 'TXP:MIN' in command:
        return 'TXP MIN'
    elif 'TXP:MAX' in command:
        return 'TXP MAX'
    
    # 提取最后一部分作为名称
    parts = command.split(':')
    if parts:
        return parts[-1].replace('?', '')
    
    return command[:20]


def _create_chart_image(loop_data: Dict[str, float], title: str, sub_title: str, config_id: str) -> Optional[BytesIO]:
    """创建单个配置的图表图像，保持原始格式"""
    try:
        # 准备数据
        sorted_loop_keys = sorted(loop_data.keys(), key=lambda x: int(x.replace('循环', '')))
        loop_values = []
        loop_labels = []
        
        for loop_key in sorted_loop_keys:
            loop_labels.append(loop_key)
            loop_values.append(loop_data[loop_key])
        
        if not loop_values:
            return None
        
        # 创建图表 - 原始尺寸
        plt.figure(figsize=(8, 5), dpi=150)
        
        # 创建条形图
        bars = plt.bar(range(len(loop_values)), loop_values, 
                      color=plt.cm.Set3(np.linspace(0, 1, len(loop_values))),
                      alpha=0.8,
                      width=0.6)
        
        # 添加数值标签
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:.3f}',
                    ha='center', va='bottom',
                    fontsize=9)
        
        # 设置图表属性
        plt.xlabel('循环轮次', fontsize=10)
        plt.ylabel('数值', fontsize=10)
        
        # 设置标题 - 按照image.png格式
        main_title = _simplify_title(title)
        plt.title(f'{main_title} ({sub_title})\n{config_id}', fontsize=12)
        
        # 设置x轴刻度
        plt.xticks(range(len(loop_labels)), loop_labels)
        
        # 添加网格
        plt.grid(True, alpha=0.3, axis='y')
        
        # 调整布局
        plt.tight_layout()
        
        # 将图表保存到BytesIO
        img_data = BytesIO()
        plt.savefig(img_data, format='png', dpi=150, 
                   bbox_inches='tight', facecolor='white')
        img_data.seek(0)
        plt.close()  # 关闭图形释放内存
        
        return img_data
        
    except Exception as e:
        print(f"❌ 创建图表图像失败: {e}")
        return None


def _simplify_title(title: str) -> str:
    """简化标题，按照image.png格式"""
    if 'BLER:UL' in title or 'FETCh:NR:BLER:UL:RESult?' in title:
        return 'UL BLER'
    elif 'BLER:DL' in title or 'FETCh:NR:BLER:DL:RESult?' in title:
        return 'DL BLER'
    elif 'TXP:AVG' in title:
        return 'TXP AVG'
    elif 'TXP:MIN' in title:
        return 'TXP MIN'
    elif 'TXP:MAX' in title:
        return 'TXP MAX'
    
    # 提取步骤内容的关键部分
    if '步骤' in title:
        # 提取步骤内容的关键部分
        if 'BLER' in title:
            if 'UL' in title:
                return 'UL BLER'
            elif 'DL' in title:
                return 'DL BLER'
        elif 'TXP' in title:
            if 'AVG' in title:
                return 'TXP AVG'
            elif 'MIN' in title:
                return 'TXP MIN'
            elif 'MAX' in title:
                return 'TXP MAX'
    
    # 截断长标题
    if len(title) > 30:
        return title[:27] + "..."
    return title


def _extract_chart_data_by_command_custom(results: List[Dict], data_extraction_config: Dict) -> Dict:
    """根据命令模式提取图表数据（自定义版本）"""
    chart_data = {}
    
    for result in results:
        execution_details = result.get('execution_details', [])
        
        for detail in execution_details:
            step_content = detail.get('content', '')
            result_str = str(detail.get('result', ''))
            loop_index = detail.get('loop_index', 1)
            
            # 检查是否匹配配置的命令模式
            for command_pattern, positions in data_extraction_config.items():
                if command_pattern in step_content:
                    # 提取数字
                    numbers = _extract_all_numbers_custom(result_str)
                    
                    if numbers:
                        # 为每个提取位置创建数据系列
                        for pos_idx, position in enumerate(positions):
                            if 1 <= position <= len(numbers):
                                value = numbers[position - 1]
                                
                                # 生成数据系列键
                                series_key = f"{command_pattern}_位置{position}"
                                
                                if series_key not in chart_data:
                                    chart_data[series_key] = {
                                        "command": command_pattern,
                                        "position": position,
                                        "data": {},
                                        "loop_indices": set()
                                    }
                                
                                # 存储循环数据
                                loop_key = f"循环{loop_index}"
                                if loop_key not in chart_data[series_key]["data"]:
                                    chart_data[series_key]["data"][loop_key] = value
                                    chart_data[series_key]["loop_indices"].add(loop_index)
    
    return chart_data


def _extract_chart_data_by_step_custom(results: List[Dict], chart_steps_filter: List[int]) -> Dict:
    """按步骤编号提取图表数据（自定义版本）"""
    chart_data = {}
    
    for result in results:
        execution_details = result.get('execution_details', [])
        
        for detail in execution_details:
            step_content = detail.get('content', '')
            result_str = str(detail.get('result', ''))
            step_num = detail.get('original_step', detail.get('step', 0))
            loop_index = detail.get('loop_index', 1)
            
            # 根据配置筛选步骤
            if step_num not in chart_steps_filter:
                continue
            
            # 只处理查询命令且有结果的数据
            if '?' in step_content and result_str and result_str != 'N/A':
                numbers = _extract_all_numbers_custom(result_str)
                
                if numbers:
                    step_key = f"步骤{step_num}"
                    loop_key = f"循环{loop_index}"
                    
                    if step_key not in chart_data:
                        chart_data[step_key] = {
                            "step_num": int(step_num),
                            "content": step_content.replace('?', '').strip(),
                            "loops": {}
                        }
                    
                    # 存储该循环的数据点
                    if loop_key not in chart_data[step_key]["loops"]:
                        chart_data[step_key]["loops"][loop_key] = {
                            "loop_index": loop_index,
                            "data_points": numbers
                        }
    
    return chart_data


def _extract_all_numbers_custom(text: str) -> List[float]:
    """从文本中提取所有数字（自定义版本）"""
    try:
        if not text:
            return []
        
        clean_text = text.strip()
        
        # 检查是否包含错误信息
        error_keywords = ["仪器通信错误", "VI_ERROR_TMO", "Timeout", "通信失败", "错误", "ERROR"]
        if any(keyword in clean_text for keyword in error_keywords):
            return []
        
        # 检查是否为空或无效响应
        if clean_text in ['', 'N/A', 'None', 'null', '[]', '{}']:
            return []
        
        # 处理逗号分隔的数组格式
        if ',' in clean_text:
            parts = [part.strip() for part in clean_text.split(',')]
            numbers = []
            for part in parts:
                try:
                    num = float(part)
                    numbers.append(num)
                except ValueError:
                    continue
            if numbers:
                return numbers
        
        # 常规提取所有数字
        pattern = r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?'
        matches = re.findall(pattern, clean_text)
        
        numbers = []
        for match in matches:
            try:
                number = float(match)
                numbers.append(number)
            except ValueError:
                continue
        
        return numbers
    except Exception as e:
        return []


# 保留独立的导出功能
if __name__ == "__main__":
    exporter = ResultExporter()
    
    # 可以传入特定的json文件路径
    # x = r"D:\feifei\2026_01_05\auto_test\log\execution_20260113_163555\souren_results_20260113_163555.json"
    success = exporter.convert_to_excel()
    
    if success:
        print("✅ Excel文件导出成功!")
    else:
        print("❌ Excel文件导出失败")