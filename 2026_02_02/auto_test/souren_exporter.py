from lib.var import *
class ResultExporter:
    """结果导出器 - 将JSON结果自动转换为Excel,使用matplotlib绘制柱状图"""
    
    def __init__(self):
        import souren_config
        
        self.log_dir = souren_config.EXECUTION_DIR
        
        self.default_row_height = getattr(souren_config, 'EXCEL_DEFAULT_ROW_HEIGHT', 13.5)
        self.default_column_width = getattr(souren_config, 'EXCEL_DEFAULT_COLUMN_WIDTH', 9)
        self.header_row_height = getattr(souren_config, 'EXCEL_HEADER_ROW_HEIGHT', 20)
        
        self.summary_column_widths = getattr(souren_config, 'EXCEL_SUMMARY_COLUMN_WIDTHS', {})
        self.details_column_widths = getattr(souren_config, 'EXCEL_DETAILS_COLUMN_WIDTHS', {})
        
        if hasattr(souren_config, 'EXCEL_CHART_COLUMN_WIDTHS'):
            self.chart_column_widths = souren_config.EXCEL_CHART_COLUMN_WIDTHS
        else:
            self.chart_column_widths = {'default': 15}
        
        self.border_colors = getattr(souren_config, 'EXCEL_BORDER_COLORS', {"default": "000000"})
        
        if hasattr(souren_config, 'generate_chart_colors'):
            self.generate_chart_colors_func = souren_config.generate_chart_colors
        else:
            self.generate_chart_colors_func = lambda count: plt.cm.tab10(np.linspace(0, 1, min(count, 10)))
        
        self.use_chart_data_extraction = getattr(souren_config, 'USE_CHART_DATA_EXTRACTION', False)
        
        if self.use_chart_data_extraction:
            self.data_extraction_config = getattr(souren_config, 'CHART_DATA_EXTRACTION_CONFIG', {})
            self.chart_steps_filter = []
            print(f"📊 使用命令模式提取图表数据")
            print(f"   配置内容: {self.data_extraction_config}")
        else:
            self.data_extraction_config = {}
            self.chart_steps_filter = getattr(souren_config, 'CHART_STEPS_FILTER', [])
            print(f"📊 使用步骤编号筛选配置")
            print(f"   配置内容: {self.chart_steps_filter}")
        
        self.header_fill = PatternFill(
            start_color=getattr(souren_config, 'EXCEL_HEADER_FILL_COLOR', "366092"), 
            end_color=getattr(souren_config, 'EXCEL_HEADER_FILL_COLOR', "366092"), 
            fill_type="solid"
        )
        self.header_font = Font(
            color=getattr(souren_config, 'EXCEL_HEADER_FONT_COLOR', "FFFFFF"), 
            bold=True, 
            size=getattr(souren_config, 'EXCEL_HEADER_FONT_SIZE', 11)
        )
        self.query_result_fill = PatternFill(
            start_color=getattr(souren_config, 'EXCEL_QUERY_RESULT_FILL', "E2F0D9"), 
            end_color=getattr(souren_config, 'EXCEL_QUERY_RESULT_FILL', "E2F0D9"), 
            fill_type="solid"
        )
        self.query_result_font = Font(
            color=getattr(souren_config, 'EXCEL_QUERY_RESULT_FONT', "006400"), 
            bold=True
        )
        
        # skipped状态的红色样式
        self.skipped_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        self.skipped_font = Font(color="FFFFFF", bold=True)
        
        self.data_font = Font(size=getattr(souren_config, 'EXCEL_DATA_FONT_SIZE', 10))
        
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        self.chart_style = {
            'figsize': (10, 6),
            'dpi': 150,
            'grid_alpha': 0.3,
            'font_size_title': 12,
            'font_size_axis': 10,
            'legend_loc': 'best',
            'bar_width': 0.6,
            'bar_alpha': 0.8,
            'colors': plt.cm.Set3(np.linspace(0, 1, 12))
        }
    
    def find_latest_json_result(self) -> Optional[str]:
        """查找最新的JSON结果文件"""
        try:
            if not os.path.exists(self.log_dir):
                return None
            
            json_files = []
            for filename in os.listdir(self.log_dir):
                if filename.startswith("souren_results_") and filename.endswith(".json"):
                    filepath = os.path.join(self.log_dir, filename)
                    if os.path.isfile(filepath):
                        mtime = os.path.getmtime(filepath)
                        json_files.append({"path": filepath, "mtime": mtime})
            
            if not json_files:
                return None
            
            json_files.sort(key=lambda x: x["mtime"], reverse=True)
            return json_files[0]["path"]
            
        except Exception:
            return None
    
    def load_json_results(self, json_file: str) -> List[Dict]:
        """加载JSON结果数据"""
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            if not content.strip():
                return []
            
            results = json.loads(content)
            return [results] if isinstance(results, dict) else results
            
        except Exception:
            return []
    
    def convert_to_excel(self, json_file: str = None) -> bool:
        """将JSON结果转换为Excel文件"""
        try:
            if json_file is None:
                json_file = self.find_latest_json_result()
                if json_file is None:
                    print("❌ 未找到JSON结果文件")
                    return False
            
            json_name = os.path.basename(json_file)
            excel_name = json_name.replace('.json', '.xlsx')
            excel_file = os.path.join(os.path.dirname(json_file), excel_name)
            
            import time
            if os.path.exists(excel_file):
                try:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                    backup_file = excel_file.replace('.xlsx', f'_{timestamp}.xlsx')
                    os.rename(excel_file, backup_file)
                    print(f"⚠️ 原文件已存在，已重命名为: {os.path.basename(backup_file)}")
                except:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    excel_name = f"souren_results_{timestamp}.xlsx"
                    excel_file = os.path.join(os.path.dirname(json_file), excel_name)
                    print(f"⚠️ 文件被锁定，使用新文件名: {excel_name}")
            
            results = self.load_json_results(json_file)
            if not results:
                print("❌ 未加载到有效的JSON结果数据")
                return False
            
            print(f"📊 开始创建Excel文件: {excel_name}")
            
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                self._create_summary_sheet(writer, results)
                self._create_execution_details_sheet(writer, results)
            
            self._apply_styles_and_charts(excel_file, results)
            
            print(f"✅ Excel文件已生成: {excel_file}")
            return True
            
        except Exception as e:
            print(f"❌ Excel文件导出失败: {e}")
            return False
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter, results: List[Dict]):
        """创建汇总表 - 优化：每个配置只统计一行，合并多次循环数据"""
        try:
            summary_data = []
            
            # 按配置分组结果
            config_groups = {}
            
            for i, result in enumerate(results):
                # 生成配置标识
                script_file = result.get('script_file', result.get('file', result.get('script_name', 'N/A')))
                
                # 从文件名中提取SCV文件名（去除路径和扩展名）
                if isinstance(script_file, str):
                    if os.path.sep in script_file:
                        # 提取文件名
                        script_file = os.path.basename(script_file)
                    # 去除.py扩展名
                    if script_file.endswith('.py'):
                        script_file = script_file[:-3]
                
                parameters = result.get('parameters', {})
                
                # 创建配置标识
                config_id = script_file
                if parameters:
                    param_parts = []
                    if 'lineLoss' in parameters:
                        param_parts.append(f"线损:{parameters['lineLoss']}")
                    if 'band' in parameters:
                        param_parts.append(f"band:{parameters['band']}")
                    if 'bw' in parameters:
                        param_parts.append(f"bw:{parameters['bw']}")
                    if 'scs' in parameters:
                        param_parts.append(f"scs:{parameters['scs']}")
                    if 'range' in parameters:
                        param_parts.append(f"range:{parameters['range']}")
                    
                    if param_parts:
                        config_id += f" ({', '.join(param_parts)})"
                
                # 添加到分组
                if config_id not in config_groups:
                    config_groups[config_id] = []
                config_groups[config_id].append(result)
            
            print(f"📊 按配置分组，共 {len(config_groups)} 个配置")
            
            # 处理每个配置组
            config_index = 1
            for config_id, config_results in config_groups.items():
                print(f"🔍 处理配置 {config_index}: {config_id}")
                
                # 初始化统计变量
                total_executed = 0
                total_passed = 0
                total_failed = 0
                total_communication_failed = 0
                total_duration = 0
                loop_count = 0
                first_result = config_results[0]
                
                # 统计所有循环的数据
                for result in config_results:
                    # 获取执行详情
                    execution_details = result.get('execution_details', [])
                    if not execution_details and 'result' in result:
                        execution_details = result.get('result', {}).get('execution_details', [])
                    
                    # 统计执行的步骤
                    executed = len([d for d in execution_details if not d.get('is_skipped', False)])
                    passed = len([d for d in execution_details if d.get('status') in ['success', 'skipped']])
                    failed = len([d for d in execution_details if d.get('status') == 'failed'])
                    comm_failed = len([d for d in execution_details if d.get('status') == 'communication_error'])
                    
                    total_executed += executed
                    total_passed += passed
                    total_failed += failed
                    total_communication_failed += comm_failed
                    
                    # 累加执行时间
                    exec_time = result.get('execution_time', result.get('duration', 0))
                    if isinstance(exec_time, (int, float)):
                        total_duration += float(exec_time)
                    
                    # 获取循环次数
                    loop_count = max(loop_count, result.get('loop_count', 1))
                
                # 计算成功率
                success_rate = round((total_passed / total_executed) * 100, 2) if total_executed > 0 else 0
                
                # 获取设备信息
                device = first_result.get('device', 'N/A')
                if isinstance(device, dict):
                    device = device.get('address', 'N/A')
                
                # 获取模式
                mode = first_result.get('mode', 'normal')
                
                # 获取状态
                all_interrupted = all(result.get('interrupted', False) for result in config_results)
                if all_interrupted:
                    status = 'interrupted'
                elif any(result.get('interrupted', False) for result in config_results):
                    status = 'partial_completed'
                else:
                    status = 'completed'
                
                # 获取总步骤数（从第一个结果获取）
                total_steps = first_result.get('total_steps', total_executed)
                
                # 获取参数信息
                parameters = first_result.get('parameters', {})
                param_info = ""
                if parameters:
                    if 'lineLoss' in parameters:
                        param_info += f"线损:{parameters['lineLoss']} "
                    if 'band' in parameters:
                        param_info += f"band:{parameters['band']} "
                    if 'bw' in parameters:
                        param_info += f"bw:{parameters['bw']} "
                    if 'scs' in parameters:
                        param_info += f"scs:{parameters['scs']} "
                    if 'range' in parameters:
                        param_info += f"range:{parameters['range']}"
                
                # 构建汇总数据
                summary = {
                    "序号": config_index,
                    "执行时间": first_result.get('timestamp_readable', first_result.get('timestamp', 'N/A')),
                    "SCV文件": script_file,  # 使用清理后的文件名
                    "设备": device,
                    "参数信息": param_info.strip() if param_info else "N/A",
                    "执行模式": mode,
                    "循环次数": loop_count,
                    "总步骤数": total_steps,
                    "已执行步骤": total_executed,
                    "通过步骤": total_passed,
                    "失败步骤": total_failed,
                    "成功率(%)": success_rate,
                    "总耗时(秒)": round(total_duration, 2),
                    "状态": status,
                    "状态消息": f"共执行{len(config_results)}个循环" + (" (部分中断)" if status == 'partial_completed' else "")
                }
                
                summary_data.append(summary)
                config_index += 1
            
            if summary_data:
                df_summary = pd.DataFrame(summary_data)
                
                # 定义列顺序 - 按照要求的顺序
                columns_order = [
                    "序号", "执行时间", "SCV文件", "设备", "参数信息",
                    "执行模式", "循环次数", "总步骤数", "已执行步骤", 
                    "通过步骤", "失败步骤", "成功率(%)", "总耗时(秒)", 
                    "状态", "状态消息"
                ]
                
                # 只保留存在的列
                columns_order = [col for col in columns_order if col in df_summary.columns]
                
                df_summary = df_summary[columns_order]
                df_summary.to_excel(writer, sheet_name='执行汇总', index=False)
                print(f"✅ 汇总表创建成功，共 {len(summary_data)} 行数据")
                
                # 调试信息：显示列宽配置
                print(f"📊 汇总表列宽配置: {self.summary_column_widths}")
                print(f"📊 汇总表实际列数: {len(columns_order)}")
                
            else:
                # 创建空的汇总表
                df_empty = pd.DataFrame({"状态": ["没有有效的执行记录"]})
                df_empty.to_excel(writer, sheet_name='执行汇总', index=False)
                print(f"⚠️ 没有找到有效的执行记录")
                
        except Exception as e:
            print(f"❌ 创建汇总表失败: {e}")
            import traceback
            traceback.print_exc()
    
    def _create_execution_details_sheet(self, writer: pd.ExcelWriter, results: List[Dict]):
        """创建详细执行记录表 - 优化：循环之间添加两行空行"""
        try:
            details_data = []
            
            print(f"📊 开始创建详细执行记录，共 {len(results)} 个结果")
            
            # 按配置分组
            config_groups = {}
            for result in results:
                script_file = result.get('script_file', result.get('file', result.get('script_name', 'N/A')))
                config_id = script_file
                if config_id not in config_groups:
                    config_groups[config_id] = []
                config_groups[config_id].append(result)
            
            # 处理每个配置
            for config_index, (config_id, config_results) in enumerate(config_groups.items(), 1):
                print(f"🔍 处理配置 {config_index}: {config_id}")
                
                # 在配置之间添加分隔行（除了第一个配置）
                if config_index > 1:
                    # 添加两行空行作为配置间隔
                    for _ in range(2):
                        details_data.append({
                            "配置": "", "执行时间": "", "步骤序号": "", "循环轮次": "", "总循环数": "",
                            "步骤内容": "", "步骤类型": "", "执行状态": "", "执行结果": "", "耗时(秒)": ""
                        })
                
                # 处理每个循环
                for loop_idx, result in enumerate(config_results, 1):
                    print(f"   处理循环 {loop_idx}/{len(config_results)}")
                    
                    # 获取执行详情
                    execution_details = result.get('execution_details', [])
                    if not execution_details and 'result' in result:
                        execution_details = result.get('result', {}).get('execution_details', [])
                    
                    print(f"     执行详情数量: {len(execution_details)}")
                    
                    if not execution_details:
                        continue
                    
                    # 在循环之间添加两行空行（除了第一个循环）
                    if loop_idx > 1:
                        for _ in range(2):
                            details_data.append({
                                "配置": "",
                                "执行时间": "",
                                "步骤序号": "",
                                "循环轮次": "",
                                "总循环数": "",
                                "步骤内容": "",
                                "步骤类型": "",
                                "执行状态": "",
                                "执行结果": "",
                                "耗时(秒)": ""
                            })
                    
                    # 处理执行详情
                    for detail in execution_details:
                        step_content = detail.get('content', '')
                        if not step_content:
                            continue
                        
                        # 跳过控制标记
                        if (step_content.startswith('for:') or 
                            step_content.startswith('if:') or 
                            step_content == 'else:'):
                            continue
                        
                        loop_index = detail.get('loop_index', 1)
                        if loop_index == 0:
                            loop_index = 1
                        
                        # 获取步骤类型
                        step_type = detail.get('type', 'Normal')
                        if step_type == 'ForLoop':
                            step_type = '循环标记'
                        elif step_type == 'IfCondition':
                            step_type = '条件判断'
                        elif step_type == 'ElseCondition':
                            step_type = '条件分支'
                        
                        # 构建详细记录
                        detail_record = {
                            "配置": f"配置{config_index}",
                            "执行时间": result.get('timestamp_readable', result.get('timestamp', 'N/A')),
                            "步骤序号": detail.get('step', 'N/A'),
                            "循环轮次": detail.get('loop_iteration', loop_idx),
                            "总循环数": detail.get('loop_count', len(config_results)),
                            "步骤内容": step_content,
                            "步骤类型": step_type,
                            "执行状态": detail.get('status', 'N/A'),
                            "执行结果": str(detail.get('result', 'N/A')),
                            "耗时(秒)": round(float(detail.get('duration', 0)), 3)
                        }
                        
                        details_data.append(detail_record)
                
                # 添加配置结束空行
                for _ in range(2):
                    details_data.append({
                        "配置": "",
                        "执行时间": "",
                        "步骤序号": "",
                        "循环轮次": "",
                        "总循环数": "",
                        "步骤内容": "",
                        "步骤类型": "",
                        "执行状态": "",
                        "执行结果": "",
                        "耗时(秒)": ""
                    })
            
            print(f"📊 详细记录总数: {len(details_data)}")
            
            if details_data:
                # 列定义
                columns = [
                    "配置", "执行时间", "步骤序号", "循环轮次", "总循环数", 
                    "步骤内容", "步骤类型", "执行状态", "执行结果", "耗时(秒)"
                ]
                
                df_details = pd.DataFrame(details_data, columns=columns)
                df_details.to_excel(writer, sheet_name='详细执行记录', index=False)
                print(f"✅ 详细执行记录表创建成功，共 {len(details_data)} 行数据")
            else:
                # 创建空的详细记录表
                df_empty = pd.DataFrame({"状态": ["没有详细的执行记录"]})
                df_empty.to_excel(writer, sheet_name='详细执行记录', index=False)
                print(f"⚠️ 没有找到详细的执行记录")
                
        except Exception as e:
            print(f"❌ 创建详细记录表失败: {e}")
            import traceback
            traceback.print_exc()
    
    def _extract_chart_data(self, results: List[Dict]) -> Dict:
        """提取图表数据 - 优化：支持多次循环数据"""
        if self.use_chart_data_extraction:
            return self._extract_chart_data_by_command_enhanced(results)
        else:
            return self._extract_chart_data_by_step_enhanced(results)
    
    def _extract_chart_data_by_command_enhanced(self, results: List[Dict]) -> Dict:
        """根据命令模式提取图表数据 - 增强版：支持多次循环"""
        chart_data = {}
        
        print(f"📊 开始根据命令模式提取数据...")
        print(f"   配置的命令模式: {list(self.data_extraction_config.keys())}")
        
        # 按配置分组
        config_groups = {}
        for result in results:
            script_file = result.get('script_file', result.get('file', result.get('script_name', 'N/A')))
            config_id = script_file
            if config_id not in config_groups:
                config_groups[config_id] = []
            config_groups[config_id].append(result)
        
        # 处理每个配置
        for config_id, config_results in config_groups.items():
            print(f"🔍 处理配置: {config_id}")
            
            # 处理每个循环
            for loop_idx, result in enumerate(config_results, 1):
                print(f"   处理循环 {loop_idx}")
                
                # 从结果中获取执行详情
                execution_details = result.get('execution_details', [])
                if not execution_details and 'result' in result:
                    execution_details = result.get('result', {}).get('execution_details', [])
                
                for detail in execution_details:
                    step_content = detail.get('content', '')
                    result_str = str(detail.get('result', ''))
                    
                    for command_pattern, positions in self.data_extraction_config.items():
                        # 使用更宽松的匹配：命令模式是否包含在步骤内容中
                        if command_pattern in step_content:
                            numbers = self._extract_all_numbers(result_str)
                            
                            if numbers:
                                for position in positions:
                                    if 1 <= position <= len(numbers):
                                        value = numbers[position - 1]
                                        
                                        # 创建唯一的数据键
                                        series_key = f"{config_id}_{command_pattern}_位置{position}"
                                        
                                        if series_key not in chart_data:
                                            chart_data[series_key] = {
                                                "config_id": config_id,
                                                "command": command_pattern,
                                                "position": position,
                                                "data": {},  # 存储每个循环的数据
                                                "loop_colors": {}  # 存储每个循环的颜色
                                            }
                                        
                                        # 存储循环数据
                                        loop_key = f"循环{loop_idx}"
                                        chart_data[series_key]["data"][loop_key] = value
                                        
                                        # 为每个循环分配不同颜色
                                        if loop_key not in chart_data[series_key]["loop_colors"]:
                                            # 使用颜色映射，确保每个循环颜色不同
                                            color_map = plt.cm.tab10(np.linspace(0, 1, 10))
                                            color_idx = (loop_idx - 1) % len(color_map)
                                            chart_data[series_key]["loop_colors"][loop_key] = color_map[color_idx]
                                        
                                        print(f"    ✅ 提取到数据: {config_id} {command_pattern} 循环{loop_idx} 位置{position} = {value:.3f}")
        
        print(f"📊 共提取了 {len(chart_data)} 个数据系列")
        return chart_data
    
    def _extract_chart_data_by_step_enhanced(self, results: List[Dict]) -> Dict:
        """按步骤编号提取图表数据 - 增强版：支持多次循环"""
        chart_data = {}
        
        # 按配置分组
        config_groups = {}
        for result in results:
            script_file = result.get('script_file', result.get('file', result.get('script_name', 'N/A')))
            config_id = script_file
            if config_id not in config_groups:
                config_groups[config_id] = []
            config_groups[config_id].append(result)
        
        # 处理每个配置
        for config_id, config_results in config_groups.items():
            print(f"🔍 处理配置: {config_id}")
            
            # 处理每个循环
            for loop_idx, result in enumerate(config_results, 1):
                print(f"   处理循环 {loop_idx}")
                
                execution_details = result.get('execution_details', [])
                if not execution_details and 'result' in result:
                    execution_details = result.get('result', {}).get('execution_details', [])
                
                for detail in execution_details:
                    step_content = detail.get('content', '')
                    result_str = str(detail.get('result', ''))
                    step_num = detail.get('original_step', detail.get('step', 0))
                    
                    if step_num not in self.chart_steps_filter:
                        continue
                    
                    if '?' in step_content and result_str and result_str != 'N/A':
                        numbers = self._extract_all_numbers(result_str)
                        
                        if numbers:
                            step_key = f"{config_id}_步骤{step_num}"
                            loop_key = f"循环{loop_idx}"
                            
                            if step_key not in chart_data:
                                chart_data[step_key] = {
                                    "config_id": config_id,
                                    "step_num": int(step_num),
                                    "content": step_content.replace('?', '').strip(),
                                    "loops": {},
                                    "loop_colors": {}
                                }
                            
                            if loop_key not in chart_data[step_key]["loops"]:
                                chart_data[step_key]["loops"][loop_key] = {
                                    "loop_index": loop_idx,
                                    "data_points": numbers
                                }
                                
                                # 为每个循环分配不同颜色
                                color_map = plt.cm.tab10(np.linspace(0, 1, 10))
                                color_idx = (loop_idx - 1) % len(color_map)
                                chart_data[step_key]["loop_colors"][loop_key] = color_map[color_idx]
        
        print(f"📊 根据步骤编号筛选了 {len(chart_data)} 个步骤的数据")
        return chart_data
    
    def _extract_all_numbers(self, text: str) -> List[float]:
        """从文本中提取所有数字，保持原值"""
        try:
            if not text or not isinstance(text, str):
                return []
            
            clean_text = str(text).strip()
            
            # 检查错误信息
            error_keywords = ["仪器通信错误", "VI_ERROR_TMO", "Timeout", "通信失败", "错误", "ERROR", "失败"]
            if any(keyword in clean_text.upper() for keyword in [k.upper() for k in error_keywords]):
                print(f"⚠️  检测到错误信息，跳过: {clean_text[:100]}")
                return []
            
            if clean_text in ['', 'N/A', 'None', 'null', '[]', '{}', '命令执行成功']:
                return []
            
            # 处理常见的返回格式
            # 1. 逗号分隔的数字
            if ',' in clean_text:
                parts = [part.strip() for part in clean_text.split(',')]
                numbers = []
                for part in parts:
                    # 尝试直接转换
                    try:
                        num = float(part)
                        numbers.append(num)
                    except ValueError:
                        # 如果失败，尝试从字符串中提取数字
                        num_match = re.search(r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?', part)
                        if num_match:
                            try:
                                numbers.append(float(num_match.group()))
                            except:
                                continue
                if numbers:
                    return numbers
            
            # 2. 单个数字
            try:
                num = float(clean_text)
                return [num]
            except ValueError:
                pass
            
            # 3. 使用正则表达式提取所有数字
            pattern = r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?'
            matches = re.findall(pattern, clean_text)
            
            numbers = []
            for match in matches:
                try:
                    number = float(match)
                    numbers.append(number)
                except ValueError:
                    continue
            
            # 4. 如果正则提取失败，尝试更宽松的匹配
            if not numbers and clean_text:
                # 查找任何看起来像数字的字符串
                num_candidates = re.findall(r'\b\d+\.?\d*\b', clean_text)
                for candidate in num_candidates:
                    try:
                        numbers.append(float(candidate))
                    except:
                        continue
            
            return numbers
        except Exception as e:
            print(f"❌ 提取数字失败: {e}, 文本: {text[:100]}..." if text else f"❌ 提取数字失败: {e}")
            return []
    
    def _create_bar_chart(self, config_id: str, command: str, position: int, 
                         loop_data: Dict[str, float], loop_colors: Dict[str, tuple]) -> BytesIO:
        """
        创建柱状图显示不同循环的数据 - 增强版：每个循环不同颜色
        """
        try:
            sorted_loop_keys = sorted(loop_data.keys(), key=lambda x: int(x.replace('循环', '')))
            loop_values = []
            loop_labels = []
            loop_colors_list = []
            
            for loop_key in sorted_loop_keys:
                loop_labels.append(loop_key)
                loop_values.append(loop_data[loop_key])
                loop_colors_list.append(loop_colors.get(loop_key, self.chart_style['colors'][0]))
            
            if not loop_values:
                return None
            
            fig, ax = plt.subplots(figsize=self.chart_style['figsize'])
            
            # 使用不同的颜色绘制每个循环的柱状图
            bars = ax.bar(range(len(loop_values)), loop_values, 
                         color=loop_colors_list, 
                         alpha=self.chart_style['bar_alpha'],
                         width=self.chart_style['bar_width'])
            
            # 在每个柱状图上添加数值标签
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{height:.3f}',
                       ha='center', va='bottom',
                       fontsize=10)
            
            ax.set_xlabel('循环轮次', fontsize=self.chart_style['font_size_axis'])
            ax.set_ylabel('数值', fontsize=self.chart_style['font_size_axis'])
            
            command_name = self._get_command_name(command)
            title = f'{config_id}\n{command_name} (位置{position})'
            ax.set_title(title, fontsize=self.chart_style['font_size_title'])
            
            ax.set_xticks(range(len(loop_labels)))
            ax.set_xticklabels(loop_labels)
            
            ax.grid(True, alpha=self.chart_style['grid_alpha'], axis='y')
            
            # 添加图例
            legend_handles = []
            for loop_key, color in zip(sorted_loop_keys, loop_colors_list):
                legend_handles.append(plt.Rectangle((0, 0), 1, 1, fc=color, label=loop_key))
            
            if legend_handles:
                ax.legend(handles=legend_handles, title='循环', loc='upper right')
            
            plt.tight_layout()
            
            img_data = BytesIO()
            plt.savefig(img_data, format='png', dpi=self.chart_style['dpi'], 
                       bbox_inches='tight', facecolor='white')
            img_data.seek(0)
            plt.close(fig)
            
            return img_data
            
        except Exception as e:
            print(f"❌ 创建柱状图失败: {e}")
            return None
    
    def _get_command_name(self, command: str) -> str:
        """从命令中提取简化的名称"""
        clean_command = command.replace('?', '').strip()
        
        if 'BLER:UL' in clean_command or 'UL:RESult' in clean_command:
            return 'UL BLER'
        elif 'BLER:DL' in clean_command or 'DL:RESult' in clean_command:
            return 'DL BLER'
        elif 'TXP:AVG' in clean_command:
            return 'TXP AVG'
        elif 'TXP:MIN' in clean_command:
            return 'TXP MIN'
        elif 'TXP:MAX' in clean_command:
            return 'TXP MAX'
        
        parts = clean_command.split(':')
        if parts:
            return parts[-1]
        
        return clean_command[:20]
    
    def _apply_styles_and_charts(self, excel_file: str, results: List[Dict]):
        """应用样式和图表到Excel文件"""
        try:
            wb = load_workbook(excel_file)
            
            if '执行汇总' in wb.sheetnames:
                self._format_summary_sheet(wb['执行汇总'])
            
            if '详细执行记录' in wb.sheetnames:
                self._format_details_sheet(wb['详细执行记录'])
            
            chart_data = self._extract_chart_data(results)
            
            if chart_data:
                self._create_chart_sheet(wb, chart_data)
            else:
                if '数据分析图表' in wb.sheetnames:
                    wb.remove(wb['数据分析图表'])
                chart_ws = wb.create_sheet(title='数据分析图表')
                chart_ws['A1'] = "数据分析图表"
                
                if self.use_chart_data_extraction:
                    chart_ws['A2'] = "⚠️ 没有找到配置命令的查询数据"
                    chart_ws['A3'] = f"配置的命令模式: {list(self.data_extraction_config.keys())}"
                    print(f"⚠️ 没有找到配置命令的查询数据")
                else:
                    chart_ws['A2'] = "⚠️ 没有找到配置步骤的查询数据"
                    chart_ws['A3'] = f"配置的步骤编号: {self.chart_steps_filter}"
            
            self._adjust_dimensions(wb)
            
            wb.save(excel_file)
            
        except Exception as e:
            print(f"❌ 应用样式和图表失败: {e}")
    
    def _create_chart_sheet(self, wb, chart_data):
        """创建图表工作表 - 增强版：显示多次循环数据"""
        try:
            if '数据分析图表' in wb.sheetnames:
                wb.remove(wb['数据分析图表'])
            
            chart_ws = wb.create_sheet(title='数据分析图表')
            
            title = "数据分析图表（基于命令模式提取）"
            config_info = f"数据提取配置: {self.data_extraction_config}"
            
            chart_ws['A1'] = title
            chart_ws['A1'].font = Font(size=14, bold=True, color=self.header_fill.start_color.rgb)
            chart_ws.merge_cells('A1:L1')
            
            chart_ws['A2'] = config_info
            chart_ws['A2'].font = Font(size=10, italic=True)
            
            print(f"📊 开始创建柱状图...")
            
            current_row = 5
            charts_created = 0
            
            for series_key, series_info in chart_data.items():
                config_id = series_info["config_id"]
                command = series_info["command"]
                position = series_info["position"]
                loop_data = series_info["data"]
                loop_colors = series_info.get("loop_colors", {})
                
                if not loop_data:
                    continue
                
                print(f"  📊 处理命令: {command}, 位置: {position}, 循环数: {len(loop_data)}")
                print(f"    循环数据: {loop_data}")
                
                # 为每个循环分配颜色（如果未分配）
                if not loop_colors:
                    color_map = plt.cm.tab10(np.linspace(0, 1, 10))
                    loop_colors = {}
                    for i, loop_key in enumerate(sorted(loop_data.keys(), key=lambda x: int(x.replace('循环', '')))):
                        color_idx = i % len(color_map)
                        loop_colors[loop_key] = color_map[color_idx]
                
                img_data = self._create_bar_chart(config_id, command, position, loop_data, loop_colors)
                
                if img_data:
                    img = ExcelImage(img_data)
                    
                    img.width = 1200
                    img.height = 600
                    
                    chart_cell = f"A{current_row}"
                    chart_ws.add_image(img, chart_cell)
                    
                    explanation_row = current_row + 20
                    chart_ws.cell(row=explanation_row, column=1, value=f"配置: {config_id}")
                    chart_ws.cell(row=explanation_row + 1, column=1, value=f"命令: {command}")
                    chart_ws.cell(row=explanation_row + 2, column=1, value=f"提取位置: {position}")
                    chart_ws.cell(row=explanation_row + 3, column=1, value=f"循环数: {len(loop_data)}")
                    
                    chart_ws.cell(row=explanation_row + 4, column=1, value="具体数值:")
                    value_row = explanation_row + 5
                    for loop_key in sorted(loop_data.keys(), key=lambda x: int(x.replace('循环', ''))):
                        value = loop_data[loop_key]
                        # 获取颜色信息
                        color_info = ""
                        if loop_key in loop_colors:
                            color = loop_colors[loop_key]
                            # 转换为RGB
                            color_rgb = f"RGB({int(color[0]*255)},{int(color[1]*255)},{int(color[2]*255)})"
                            color_info = f" [颜色: {color_rgb}]"
                        
                        chart_ws.cell(row=value_row, column=1, value=f"{loop_key}: {value:.3f}{color_info}")
                        value_row += 1
                    
                    current_row = value_row + 3
                    charts_created += 1
                else:
                    print(f"  ❌ 创建图表失败: {command}")
            
            if charts_created == 0:
                chart_ws['A5'] = "⚠️ 没有足够的数据来创建图表"
                print("⚠️ 没有足够的数据来创建图表")
            else:
                print(f"✅ 成功创建了 {charts_created} 个柱状图并插入到Excel")
            
            for col_letter, width in self.chart_column_widths.items():
                if col_letter == 'default':
                    for col in range(1, chart_ws.max_column + 1):
                        col_letter = get_column_letter(col)
                        if col_letter not in self.chart_column_widths:
                            chart_ws.column_dimensions[col_letter].width = width
                else:
                    if col_letter in chart_ws.column_dimensions:
                        chart_ws.column_dimensions[col_letter].width = width
            
        except Exception as e:
            print(f"❌ 创建图表工作表失败: {e}")
    
    def _format_summary_sheet(self, ws):
        """格式化汇总表"""
        try:
            # 设置表头样式
            for cell in ws[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.border = self._create_border(self.border_colors.get("default", "000000"))
            
            # 设置数据行样式
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = self._create_border(self.border_colors.get("default", "000000"))
                    cell.font = self.data_font
                    
                    col_idx = cell.column
                    col_letter = get_column_letter(col_idx)
                    
                    # 根据列类型设置对齐方式
                    if col_letter in ["A", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
                        cell.alignment = self.center_align
                    elif col_letter in ["B", "C", "D", "E", "O"]:
                        cell.alignment = self.left_align
                
                ws.row_dimensions[row[0].row].height = self.default_row_height
            
        except Exception as e:
            print(f"❌ 格式化汇总表失败: {e}")
    
    def _format_details_sheet(self, ws):
        """格式化详细执行记录表 - 修复：只将skipped状态设为红色"""
        try:
            # 设置表头样式
            for cell in ws[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.border = self._create_border(self.border_colors.get("default", "000000"))
            
            ws.row_dimensions[1].height = self.header_row_height
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                # 跳过空行
                if not any(cell.value for cell in row):
                    continue
                
                # 检查是否为skipped状态（第8列是执行状态）
                status_cell = row[7]  # 执行状态列索引为7（0-based）
                is_skipped = False
                if status_cell.value:
                    status_value = str(status_cell.value).strip().lower()
                    is_skipped = status_value == "skipped"
                
                step_content_cell = row[5]  # 步骤内容列索引为5
                is_query = step_content_cell.value and '?' in str(step_content_cell.value)
                
                # 先应用基本样式
                for cell in row:
                    cell.border = self._create_border(self.border_colors.get("default", "000000"))
                    cell.font = self.data_font
                    
                    col_idx = cell.column
                    if col_idx in [1, 3, 4, 5, 7, 8, 10]:  # 中心对齐的列
                        cell.alignment = self.center_align
                    elif col_idx == 2:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = self.left_align
                
                # 只将skipped状态设为红色
                if is_skipped:
                    for cell in row:
                        cell.fill = self.skipped_fill
                        cell.font = self.skipped_font
                # 如果是查询命令，使用绿色背景（但skipped优先）
                elif is_query and not is_skipped:
                    for cell in row:
                        cell.fill = self.query_result_fill
                        cell.font = self.query_result_font
                
                ws.row_dimensions[row_idx].height = self.default_row_height
            
        except Exception as e:
            print(f"❌ 格式化详细记录表失败: {e}")
    
    def _create_border(self, color: str) -> Border:
        """创建边框"""
        side = Side(style='thin', color=color)
        return Border(left=side, right=side, top=side, bottom=side)
    
    def _adjust_dimensions(self, wb):
        """调整列宽和行高 - 确保列宽配置生效"""
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 设置行高
                for row in range(1, ws.max_row + 1):
                    if row == 1:
                        ws.row_dimensions[row].height = self.header_row_height
                    else:
                        # 保持默认行高
                        if ws.row_dimensions[row].height is None:
                            ws.row_dimensions[row].height = self.default_row_height
                
                if sheet_name == '执行汇总':
                    print(f"🔧 应用执行汇总表列宽配置:")
                    # 先设置默认列宽
                    for col in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(col)
                        ws.column_dimensions[col_letter].width = self.default_column_width
                    
                    # 应用特殊列宽配置
                    for col_letter, width in self.summary_column_widths.items():
                        if col_letter in ws.column_dimensions:
                            ws.column_dimensions[col_letter].width = width
                            print(f"   列 {col_letter}: {width}")
                        else:
                            print(f"   ⚠️ 列 {col_letter} 不存在于工作表中")
                            
                elif sheet_name == '详细执行记录':
                    # 先设置默认列宽
                    for col in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(col)
                        ws.column_dimensions[col_letter].width = self.default_column_width
                    
                    # 应用特殊列宽配置
                    for col_letter, width in self.details_column_widths.items():
                        if col_letter in ws.column_dimensions:
                            ws.column_dimensions[col_letter].width = width
                            
                elif sheet_name == '数据分析图表':
                    for col_letter, width in self.chart_column_widths.items():
                        if col_letter == 'default':
                            for col in range(1, ws.max_column + 1):
                                col_letter = get_column_letter(col)
                                if col_letter not in self.chart_column_widths:
                                    ws.column_dimensions[col_letter].width = width
                        elif col_letter in ws.column_dimensions:
                            ws.column_dimensions[col_letter].width = width
                            
                else:
                    # 其他工作表使用默认列宽
                    for col in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(col)
                        ws.column_dimensions[col_letter].width = self.default_column_width
                
                # 设置自动筛选（除了图表工作表）
                if sheet_name != '数据分析图表' and ws.max_row > 1:
                    try:
                        ws.auto_filter.ref = ws.dimensions
                    except:
                        pass
            
            print(f"✅ 已应用所有Excel导出配置")
            
        except Exception as e:
            print(f"❌ 调整列宽行高失败: {e}")

def create_summary_excel(main_execution_dir: str, all_results: List[Dict]) -> Optional[str]:
    """
    创建汇总Excel文件 - 优化版：包含横向对比图表
    """
    try:
        print(f"\n{'='*60}")
        print("📊 开始创建图表汇总Excel文件...")
        print(f"{'='*60}")
        
        if not main_execution_dir or not os.path.exists(main_execution_dir):
            print("❌ 主执行目录不存在，无法创建汇总文件")
            return None
        
        if not all_results or len(all_results) == 0:
            print("❌ 没有可汇总的结果数据")
            return None
        
        try:
            from souren_config import (
                USE_CHART_DATA_EXTRACTION,
                CHART_DATA_EXTRACTION_CONFIG,
                CHART_STEPS_FILTER,
                PYTHON_SCRIPT_NAME,
                LOOP_COUNT
            )
        except ImportError as e:
            print(f"❌ 导入配置失败: {e}")
            return None
        
        print(f"📊 配置信息:")
        print(f"   USE_CHART_DATA_EXTRACTION: {USE_CHART_DATA_EXTRACTION}")
        print(f"   PYTHON_SCRIPT_NAME 数量: {len(PYTHON_SCRIPT_NAME)}")
        print(f"   LOOP_COUNT: {LOOP_COUNT}")
        
        # 1. 先处理 all_results 中的结果数据
        all_chart_data = {}
        
        for result in all_results:
            script_name = result.get('script_name', '未知脚本')
            parameters = result.get('parameters', {})
            script_dir = result.get('script_directory', '')
            interrupted = result.get('interrupted', False)
            
            # 生成配置ID，与目录名保持一致
            script_base = os.path.splitext(script_name)[0]
            config_id = script_base
            
            if parameters:
                param_str = ""
                if 'lineLoss' in parameters:
                    lineLoss_value = parameters['lineLoss']
                    if isinstance(lineLoss_value, (int, float)):
                        lineLoss_int = int(lineLoss_value)
                        param_str += f"_ll{lineLoss_int}"
                    else:
                        try:
                            lineLoss_int = int(float(lineLoss_value))
                            param_str += f"_ll{lineLoss_int}"
                        except:
                            param_str += f"_ll{lineLoss_value}"
                if 'band' in parameters:
                    param_str += f"_b{parameters['band']}"
                if 'bw' in parameters:
                    param_str += f"_bw{parameters['bw']}"
                if 'scs' in parameters:
                    param_str += f"_scs{parameters['scs']}"
                if 'range' in parameters:
                    param_str += f"_{parameters['range']}"
                
                config_id = f"{script_base}{param_str}"
            
            if interrupted:
                config_id += "_中断"
            
            print(f"🔍 处理配置: {config_id}" + (" (中断)" if interrupted else ""))
            
            # 尝试从 result 中直接提取数据
            chart_data = {}
            if USE_CHART_DATA_EXTRACTION:
                chart_data = _extract_chart_data_by_command_custom(
                    [result], 
                    CHART_DATA_EXTRACTION_CONFIG
                )
            else:
                chart_data = _extract_chart_data_by_step_custom(
                    [result],
                    CHART_STEPS_FILTER
                )
            
            if chart_data:
                all_chart_data[config_id] = {
                    'config_info': {
                        'script_name': script_name,
                        'parameters': parameters,
                        'script_dir': script_dir,
                        'loop_count': result.get('loop_count', 1),
                        'result_data': result,
                        'interrupted': interrupted
                    },
                    'chart_data': chart_data,
                    'source': 'direct_result',
                    'interrupted': interrupted
                }
                print(f"✅ 从结果数据中直接提取到图表数据: {len(chart_data)} 个数据项")
                continue
            
            # 2. 如果直接提取失败，尝试从JSON文件中提取
            if script_dir and os.path.exists(script_dir):
                json_files = []
                for file in os.listdir(script_dir):
                    if file.endswith(".json") and not file.startswith("summary"):
                        json_file = os.path.join(script_dir, file)
                        json_files.append(json_file)
                
                if json_files:
                    json_file = json_files[-1]  # 取最新的JSON文件
                    
                    try:
                        with open(json_file, 'r', encoding='utf-8') as f:
                            content = f.read()
                        
                        if not content.strip():
                            print(f"❌ JSON文件为空: {json_file}")
                            continue
                        
                        results_data = json.loads(content)
                        if isinstance(results_data, dict):
                            results_data = [results_data]
                        
                        print(f"✅ 从JSON文件加载数据: {os.path.basename(json_file)}")
                        
                        chart_data = {}
                        if USE_CHART_DATA_EXTRACTION:
                            chart_data = _extract_chart_data_by_command_custom(
                                results_data, 
                                CHART_DATA_EXTRACTION_CONFIG
                            )
                        else:
                            chart_data = _extract_chart_data_by_step_custom(
                                results_data,
                                CHART_STEPS_FILTER
                            )
                        
                        if chart_data:
                            all_chart_data[config_id] = {
                                'config_info': {
                                    'script_name': script_name,
                                    'parameters': parameters,
                                    'script_dir': script_dir,
                                    'loop_count': result.get('loop_count', 1),
                                    'result_data': result,
                                    'interrupted': interrupted
                                },
                                'chart_data': chart_data,
                                'json_file': json_file,
                                'source': 'json_file',
                                'interrupted': interrupted
                            }
                            print(f"✅ 从JSON文件中提取到图表数据: {len(chart_data)} 个数据项")
                        else:
                            print(f"⚠️  未提取到 {config_id} 的图表数据")
                            
                    except Exception as e:
                        print(f"❌ 处理JSON文件失败: {e}")
                        continue
                else:
                    print(f"❌ 在目录 {script_dir} 中未找到JSON结果文件")
            else:
                print(f"❌ 脚本目录不存在: {script_dir}")
        
        if not all_chart_data:
            print("❌ 没有提取到任何图表数据")
            # 尝试从主执行目录的子目录中查找所有JSON文件
            print("🔄 尝试从所有子目录中查找JSON文件...")
            all_chart_data = _find_all_json_files_in_subdirs(main_execution_dir, 
                                                           USE_CHART_DATA_EXTRACTION,
                                                           CHART_DATA_EXTRACTION_CONFIG,
                                                           CHART_STEPS_FILTER)
        
        if not all_chart_data:
            print("❌ 仍然没有提取到任何图表数据")
            return None
        
        print(f"📊 成功提取到 {len(all_chart_data)} 个配置的图表数据")
        
        # 3. 创建汇总Excel文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        summary_excel_name = f"summary_charts_{timestamp}.xlsx"
        summary_excel_path = os.path.join(main_execution_dir, summary_excel_name)
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as ExcelImage
        
        wb = Workbook()
        ws = wb.active
        ws.title = "图表汇总"
        
        ws['A1'] = "测试数据图表汇总"
        ws['A1'].font = Font(size=16, bold=True, color="366092")
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        ws['A3'] = f"配置数量: {len(all_chart_data)}"
        
        interrupted_count = sum(1 for data_info in all_chart_data.values() if data_info.get('interrupted', False))
        if interrupted_count > 0:
            ws['B3'] = f"中断配置: {interrupted_count}"
            ws['B3'].font = Font(size=10, color="FF0000", bold=True)
        
        current_row = 5
        
        charts_created = 0
        
        for config_id, data_info in all_chart_data.items():
            config_info = data_info['config_info']
            chart_data = data_info['chart_data']
            interrupted = data_info.get('interrupted', False)
            script_name = config_info['script_name']
            parameters = config_info['parameters']
            
            title_cell = ws.cell(row=current_row, column=1, value=f"配置: {config_id}")
            if interrupted:
                title_cell.font = Font(size=12, bold=True, color="FF0000")
                title_cell.value = f"配置: {config_id} ⚠️ (中断)"
            else:
                title_cell.font = Font(size=12, bold=True)
            current_row += 1
            
            if parameters:
                param_info = ""
                if 'lineLoss' in parameters:
                    param_info += f"线损:{parameters['lineLoss']} "
                if 'band' in parameters:
                    param_info += f"band:{parameters['band']} "
                if 'bw' in parameters:
                    param_info += f"bw:{parameters['bw']} "
                if 'scs' in parameters:
                    param_info += f"scs:{parameters['scs']} "
                if 'range' in parameters:
                    param_info += f"range:{parameters['range']}"
                
                ws.cell(row=current_row, column=1, value=f"参数: {param_info}")
                current_row += 1
            
            if interrupted:
                ws.cell(row=current_row, column=1, value="⚠️ 注意: 此配置的执行被中断，数据可能不完整")
                ws.cell(row=current_row, column=1).font = Font(size=10, color="FF0000", italic=True)
                current_row += 1
            
            # 处理每个图表数据
            for data_key, data_item in chart_data.items():
                if 'data' in data_item:
                    loop_data = data_item['data']
                    command = data_item['command']
                    position = data_item['position']
                    
                    if not loop_data:
                        continue
                    
                    print(f"  📊 处理命令: {command}, 位置: {position}")
                    print(f"    循环数据: {loop_data}")
                    
                    img_data = _create_chart_image(
                        loop_data, 
                        command, 
                        position,
                        config_id
                    )
                    
                    if img_data:
                        img = ExcelImage(img_data)
                        img.width = 600
                        img.height = 300
                        
                        chart_cell = f"A{current_row}"
                        ws.add_image(img, chart_cell)
                        
                        ws.cell(row=current_row + 15, column=1, value=f"命令: {command}")
                        ws.cell(row=current_row + 16, column=1, value=f"提取位置: {position}")
                        
                        value_row = current_row + 17
                        ws.cell(row=value_row, column=1, value="循环数据:")
                        for loop_key, value in sorted(loop_data.items()):
                            ws.cell(row=value_row, column=2, value=f"{loop_key}: {value:.3f}")
                            value_row += 1
                        
                        current_row = value_row + 2
                        charts_created += 1
                
                elif 'loops' in data_item:
                    loops_data = data_item['loops']
                    step_num = data_item.get('step_num', '?')
                    step_content = data_item.get('content', '')
                    
                    loop_data = {}
                    for loop_key, loop_info in loops_data.items():
                        data_points = loop_info.get('data_points', [])
                        if data_points:
                            loop_data[loop_key] = data_points[0]
                    
                    if loop_data:
                        img_data = _create_chart_image(
                            loop_data, 
                            step_content, 
                            step_num,
                            config_id
                        )
                        
                        if img_data:
                            img = ExcelImage(img_data)
                            img.width = 600
                            img.height = 300
                            
                            chart_cell = f"A{current_row}"
                            ws.add_image(img, chart_cell)
                            
                            ws.cell(row=current_row + 15, column=1, value=f"步骤: {step_num}")
                            ws.cell(row=current_row + 16, column=1, value=f"内容: {step_content}")
                            
                            value_row = current_row + 17
                            ws.cell(row=value_row, column=1, value="循环数据:")
                            for loop_key, value in sorted(loop_data.items()):
                                ws.cell(row=value_row, column=2, value=f"{loop_key}: {value:.3f}")
                                value_row += 1
                            
                            current_row = value_row + 2
                            charts_created += 1
            
            current_row += 2
        
        if charts_created == 0:
            print("⚠️ 没有创建任何图表")
            ws.cell(row=current_row, column=1, value="⚠️ 没有足够的数据来创建图表")
        
        # 创建横向对比图表
        comparison_ws = wb.create_sheet(title="横向对比")
        comparison_created = _create_comparison_charts(comparison_ws, all_chart_data, 
                                                     USE_CHART_DATA_EXTRACTION,
                                                     CHART_DATA_EXTRACTION_CONFIG,
                                                     PYTHON_SCRIPT_NAME,
                                                     LOOP_COUNT)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
        comparison_ws.column_dimensions['A'].width = 30
        comparison_ws.column_dimensions['B'].width = 20
        comparison_ws.column_dimensions['C'].width = 15
        
        wb.save(summary_excel_path)
        
        print(f"\n{'='*60}")
        print(f"✅ 图表汇总Excel文件创建成功!")
        print(f"📁 文件路径: {summary_excel_path}")
        print(f"📊 图表汇总: {charts_created} 个图表")
        if comparison_created:
            print(f"📈 横向对比: 已创建")
        if interrupted_count > 0:
            print(f"⚠️  包含 {interrupted_count} 个中断配置的数据")
        print(f"{'='*60}")
        
        return summary_excel_path
        
    except Exception as e:
        print(f"❌ 创建图表汇总Excel文件失败: {e}")
        import traceback
        traceback.print_exc()
        return None

def _find_all_json_files_in_subdirs(main_execution_dir, use_chart_data_extraction, 
                                   data_extraction_config, chart_steps_filter):
    """从所有子目录中查找JSON文件并提取数据"""
    all_chart_data = {}
    
    if not os.path.exists(main_execution_dir):
        return all_chart_data
    
    subdirs = [d for d in os.listdir(main_execution_dir) 
              if os.path.isdir(os.path.join(main_execution_dir, d))]
    
    print(f"🔍 在主目录中找到 {len(subdirs)} 个子目录")
    
    for subdir in subdirs:
        subdir_path = os.path.join(main_execution_dir, subdir)
        config_id = subdir  # 使用目录名作为配置ID
        
        print(f"🔍 搜索子目录: {subdir}")
        
        json_files = []
        for file in os.listdir(subdir_path):
            if file.endswith(".json") and not file.startswith("summary"):
                json_file = os.path.join(subdir_path, file)
                json_files.append(json_file)
        
        if not json_files:
            continue
        
        json_file = json_files[-1]  # 取最新的JSON文件
        
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            if not content.strip():
                continue
            
            results_data = json.loads(content)
            if isinstance(results_data, dict):
                results_data = [results_data]
            
            chart_data = {}
            if use_chart_data_extraction:
                chart_data = _extract_chart_data_by_command_custom(
                    results_data, 
                    data_extraction_config
                )
            else:
                chart_data = _extract_chart_data_by_step_custom(
                    results_data,
                    chart_steps_filter
                )
            
            if chart_data:
                all_chart_data[config_id] = {
                    'config_info': {
                        'script_name': f"{subdir}.py",
                        'parameters': {},  # 从目录名解析参数
                        'script_dir': subdir_path,
                        'loop_count': 1,
                        'result_data': None,
                        'interrupted': False
                    },
                    'chart_data': chart_data,
                    'json_file': json_file,
                    'source': 'subdir_json',
                    'interrupted': False
                }
                print(f"✅ 从子目录 {subdir} 中提取到图表数据: {len(chart_data)} 个数据项")
                
        except Exception as e:
            print(f"❌ 处理子目录 {subdir} 失败: {e}")
            continue
    
    return all_chart_data


def _create_comparison_charts(ws, all_chart_data, use_chart_data_extraction,
                            data_extraction_config, python_script_name, loop_count):
    """创建横向对比图表"""
    try:
        ws['A1'] = "配置横向对比图表"
        ws['A1'].font = Font(size=16, bold=True, color="366092")
        ws.merge_cells('A1:H1')
        
        ws['A2'] = "将所有配置的相同指标放在一起对比，一眼看出哪个配置出问题"
        ws['A2'].font = Font(size=10, italic=True)
        
        comparison_current_row = 5
        
        if not use_chart_data_extraction or not data_extraction_config:
            return 0
        
        charts_created = 0
        
        for command_pattern, positions in data_extraction_config.items():
            for position in positions:
                comparison_data = {}
                
                for config_id, data_info in all_chart_data.items():
                    chart_data = data_info['chart_data']
                    config_info = data_info['config_info']
                    interrupted = data_info.get('interrupted', False)
                    
                    series_key = f"{command_pattern}_位置{position}"
                    if series_key in chart_data:
                        loop_data = chart_data[series_key].get("data", {})
                        comparison_data[config_id] = {
                            'config_info': config_info,
                            'loop_data': loop_data,
                            'interrupted': interrupted
                        }
                
                if len(comparison_data) > 1:
                    print(f"  📊 创建横向对比图表: {command_pattern} 位置{position}")
                    print(f"     涉及配置: {list(comparison_data.keys())}")
                    
                    comparison_img_data = _create_comparison_chart_image(
                        comparison_data,
                        command_pattern,
                        position,
                        len(python_script_name),
                        loop_count
                    )
                    
                    if comparison_img_data:
                        img = ExcelImage(comparison_img_data)
                        
                        config_count = len(comparison_data)
                        img_width = min(2000, 1000 + config_count * 60)
                        img_height = min(1200, 600 + config_count * 30)
                        img.width = img_width
                        img.height = img_height
                        
                        chart_cell = f"A{comparison_current_row}"
                        ws.add_image(img, chart_cell)
                        
                        info_row = comparison_current_row + int(img_height / 15) + 5
                        ws.cell(row=info_row, column=1, 
                                value=f"命令: {_simplify_command_name(command_pattern)}")
                        ws.cell(row=info_row + 1, column=1, 
                                value=f"提取位置: {position}")
                        ws.cell(row=info_row + 2, column=1, 
                                value=f"配置数量: {len(comparison_data)}")
                        
                        interrupted_in_comparison = sum(1 for data in comparison_data.values() 
                                                       if data.get('interrupted', False))
                        if interrupted_in_comparison > 0:
                            ws.cell(row=info_row + 3, column=1, 
                                    value=f"⚠️ 包含 {interrupted_in_comparison} 个中断的配置")
                            ws.cell(row=info_row + 3, column=1).font = Font(size=10, color="FF0000", bold=True)
                        
                        comparison_current_row = info_row + 10
                        charts_created += 1
        
        return charts_created
        
    except Exception as e:
        print(f"❌ 创建横向对比图表失败: {e}")
        return 0
    
def _create_comparison_chart_image(comparison_data: Dict, command: str, position: int, 
                                  config_count: int, loop_count: int) -> Optional[BytesIO]:
    """创建横向对比图表图像，显示多个配置的循环数据对比"""
    try:
        config_ids = []
        loop_labels = []
        
        all_loops = set()
        for config_id, data in comparison_data.items():
            loop_data = data['loop_data']
            all_loops.update(loop_data.keys())
        
        loop_labels = sorted(list(all_loops), key=lambda x: int(x.replace('循环', '')))
        
        loop_series = {}
        for loop_label in loop_labels:
            loop_series[loop_label] = []
        
        for config_id, data in comparison_data.items():
            config_ids.append(config_id)
            loop_data = data['loop_data']
            interrupted = data.get('interrupted', False)
            
            display_id = config_id
            if interrupted:
                display_id = f"{config_id}⚠️"
            config_ids[-1] = display_id
            
            for loop_label in loop_labels:
                value = loop_data.get(loop_label, 0)
                loop_series[loop_label].append(value)
        
        if not config_ids or not loop_labels:
            return None
        
        base_width = 14
        base_height = 9
        width_multiplier = 0.6
        height_multiplier = 0.4
        
        dynamic_width = base_width + (config_count * width_multiplier)
        dynamic_height = base_height + (loop_count * height_multiplier)
        
        max_width = 30
        max_height = 20
        min_width = 14
        min_height = 9
        
        fig_width = min(max_width, max(min_width, dynamic_width))
        fig_height = min(max_height, max(min_height, dynamic_height))
        
        print(f"  📐 图表尺寸: {fig_width:.1f} x {fig_height:.1f} (配置: {config_count}, 循环: {loop_count})")
        
        fig, ax = plt.subplots(figsize=(fig_width, fig_height), dpi=180)
        
        x = np.arange(len(config_ids))
        
        bar_width = 0.9 / max(len(loop_labels), 1)
        bar_width = min(0.35, bar_width)
        
        group_spacing = 1.3
        x_positions = x * group_spacing
        
        colors = plt.cm.Set3(np.linspace(0, 1, len(loop_labels)))
        
        config_bars = {}
        
        for i, (loop_label, color) in enumerate(zip(loop_labels, colors)):
            values = loop_series[loop_label]
            
            bar_positions = x_positions + i * bar_width - (len(loop_labels)-1) * bar_width / 2
            bars = ax.bar(bar_positions, values, bar_width, label=loop_label, color=color, alpha=0.85)
            
            for j, bar in enumerate(bars):
                config_id = config_ids[j]
                if config_id not in config_bars:
                    config_bars[config_id] = []
                config_bars[config_id].append(bar)
            
            for bar in bars:
                height = bar.get_height()
                if height > 0:
                    label_fontsize = 10 if config_count <= 15 else 8
                    ax.text(bar.get_x() + bar.get_width()/2., height,
                           f'{height:.3f}',
                           ha='center', va='bottom',
                           fontsize=label_fontsize,
                           fontweight='bold',
                           color='black',
                           bbox=dict(boxstyle="round,pad=0.1", facecolor="white", edgecolor="none", alpha=0.7))
        
        ax.set_xlabel('配置', fontsize=13, fontweight='bold')
        ax.set_ylabel('数值', fontsize=13, fontweight='bold')
        
        ax.set_xticks(x_positions)
        
        if config_count > 15:
            rotation_angle = 60
            font_size = 6
        elif config_count > 8:
            rotation_angle = 45
            font_size = 8
        else:
            rotation_angle = 0
            font_size = 9
            
        ax.set_xticklabels(config_ids, rotation=rotation_angle, 
                          ha='right' if rotation_angle > 0 else 'center', 
                          fontsize=font_size, fontweight='bold')
        
        ax.tick_params(axis='y', labelsize=11)
        
        ax.grid(True, alpha=0.4, axis='y', linestyle='--', linewidth=0.5)
        
        if len(loop_labels) > 0:
            legend_fontsize = 10 if loop_count <= 10 else 9
            legend_title_fontsize = 11 if loop_count <= 10 else 10
            
            ax.legend(title='循环轮次', loc='upper left', bbox_to_anchor=(1.02, 1), 
                     fontsize=legend_fontsize, title_fontsize=legend_title_fontsize, 
                     framealpha=0.9, frameon=True, edgecolor='gray')
        
        right_margin = 0.78 if loop_count <= 10 else 0.82
        bottom_margin = 0.18 if config_count <= 10 else 0.25
        top_margin = 0.95
        
        plt.subplots_adjust(right=right_margin, top=top_margin, bottom=bottom_margin, left=0.12)
        
        img_data = BytesIO()
        plt.savefig(img_data, format='png', dpi=180, 
                   bbox_inches='tight', facecolor='white', pad_inches=0.3)
        img_data.seek(0)
        plt.close(fig)
        
        return img_data
        
    except Exception as e:
        print(f"❌ 创建横向对比图表失败: {e}")
        return None

def _simplify_command_name(command: str) -> str:
    """简化命令名称"""
    if 'BLER:UL' in command or 'UL:RESult' in command:
        return 'UL BLER'
    elif 'BLER:DL' in command or 'DL:RESult' in command:
        return 'DL BLER'
    elif 'TXP:AVG' in command:
        return 'TXP AVG'
    elif 'TXP:MIN' in command:
        return 'TXP MIN'
    elif 'TXP:MAX' in command:
        return 'TXP MAX'
    
    parts = command.split(':')
    if parts:
        return parts[-1].replace('?', '')
    
    return command[:20]

def _create_chart_image(loop_data: Dict[str, float], title: str, sub_title: str, config_id: str) -> Optional[BytesIO]:
    """创建单个配置的图表图像，保持原始格式"""
    try:
        sorted_loop_keys = sorted(loop_data.keys(), key=lambda x: int(x.replace('循环', '')))
        loop_values = []
        loop_labels = []
        
        for loop_key in sorted_loop_keys:
            loop_labels.append(loop_key)
            loop_values.append(loop_data[loop_key])
        
        if not loop_values:
            return None
        
        plt.figure(figsize=(8, 5), dpi=150)
        
        # 为每个循环分配不同颜色
        colors = plt.cm.tab10(np.linspace(0, 1, len(loop_values)))
        
        bars = plt.bar(range(len(loop_values)), loop_values, 
                      color=colors,
                      alpha=0.8,
                      width=0.6)
        
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:.3f}',
                    ha='center', va='bottom',
                    fontsize=9)
        
        plt.xlabel('循环轮次', fontsize=10)
        plt.ylabel('数值', fontsize=10)
        
        main_title = _simplify_title(title)
        plt.title(f'{main_title} ({sub_title})\n{config_id}', fontsize=12)
        
        plt.xticks(range(len(loop_labels)), loop_labels)
        
        plt.grid(True, alpha=0.3, axis='y')
        
        plt.tight_layout()
        
        img_data = BytesIO()
        plt.savefig(img_data, format='png', dpi=150, 
                   bbox_inches='tight', facecolor='white')
        img_data.seek(0)
        plt.close()
        
        return img_data
        
    except Exception as e:
        print(f"❌ 创建图表图像失败: {e}")
        return None

def _simplify_title(title: str) -> str:
    if 'BLER:UL' in title or 'FETCh:NR:BLER:UL:RESult?' in title:
        return 'UL BLER'
    elif 'BLER:DL' in title or 'FETCh:NR:BLER:DL:RESult?' in title:
        return 'DL BLER'
    elif 'TXP:AVG' in title:
        return 'TXP AVG'
    elif 'TXP:MIN' in title:
        return 'TXP MIN'
    elif 'TXP:MAX' in title:
        return 'TXP MAX'
    
    if '步骤' in title:
        if 'BLER' in title:
            if 'UL' in title:
                return 'UL BLER'
            elif 'DL' in title:
                return 'DL BLER'
        elif 'TXP' in title:
            if 'AVG' in title:
                return 'TXP AVG'
            elif 'MIN' in title:
                return 'TXP MIN'
            elif 'MAX' in title:
                return 'TXP MAX'
    
    if len(title) > 30:
        return title[:27] + "..."
    return title

def _extract_chart_data_by_command_custom(results: List[Dict], data_extraction_config: Dict) -> Dict:
    """根据命令模式提取图表数据（自定义版本）"""
    chart_data = {}
    
    print(f"📊 开始根据命令模式提取数据...")
    print(f"   配置的命令模式: {list(data_extraction_config.keys())}")
    
    for result in results:
        # 从结果中获取执行详情，尝试不同路径
        execution_details = result.get('execution_details', [])
        
        # 如果一级没有，尝试从 result.result 中获取
        if not execution_details:
            execution_details = result.get('result', {}).get('execution_details', [])
        
        print(f"🔍 处理结果，执行详情数量: {len(execution_details)}")
        
        for detail in execution_details:
            step_content = detail.get('content', '')
            result_str = str(detail.get('result', ''))
            loop_index = detail.get('loop_iteration', 1)  # 使用loop_iteration
            
            # 调试信息
            if 'FETCh:' in step_content and 'RESult?' in step_content:
                print(f"   处理命令: {step_content}")
                print(f"   结果字符串: {result_str[:100]}...")
                print(f"   循环索引: {loop_index}")
            
            for command_pattern, positions in data_extraction_config.items():
                # 使用更宽松的匹配：命令模式是否包含在步骤内容中
                if command_pattern in step_content:
                    numbers = _extract_all_numbers_custom(result_str)
                    
                    if numbers:
                        print(f"   提取到数字: {numbers}")
                        print(f"   数字长度: {len(numbers)}")
                        
                        for pos_idx, position in enumerate(positions):
                            if 1 <= position <= len(numbers):
                                value = numbers[position - 1]
                                
                                series_key = f"{command_pattern}_位置{position}"
                                
                                if series_key not in chart_data:
                                    chart_data[series_key] = {
                                        "command": command_pattern,
                                        "position": position,
                                        "data": {},
                                        "loop_colors": {}
                                    }
                                
                                loop_key = f"循环{loop_index}"
                                if loop_key not in chart_data[series_key]["data"]:
                                    chart_data[series_key]["data"][loop_key] = value
                                    
                                    # 为每个循环分配颜色
                                    if loop_key not in chart_data[series_key]["loop_colors"]:
                                        color_map = plt.cm.tab10(np.linspace(0, 1, 10))
                                        color_idx = (loop_index - 1) % len(color_map)
                                        chart_data[series_key]["loop_colors"][loop_key] = color_map[color_idx]
                                
                                print(f"    ✅ 提取到数据: {command_pattern} 循环{loop_index} 位置{position} = {value:.3f}")
                            else:
                                print(f"    ❌ 位置{position}超出范围，可用数字长度: {len(numbers)}")
                    else:
                        print(f"    ❌ 未提取到数字: {result_str}")
    
    print(f"📊 共提取了 {len(chart_data)} 个数据系列")
    return chart_data

def _extract_chart_data_by_step_custom(results: List[Dict], chart_steps_filter: List[int]) -> Dict:
    """按步骤编号提取图表数据（自定义版本）"""
    chart_data = {}
    
    for result in results:
        execution_details = result.get('execution_details', [])
        
        for detail in execution_details:
            step_content = detail.get('content', '')
            result_str = str(detail.get('result', ''))
            step_num = detail.get('original_step', detail.get('step', 0))
            loop_index = detail.get('loop_iteration', 1)
            
            if step_num not in chart_steps_filter:
                continue
            
            if '?' in step_content and result_str and result_str != 'N/A':
                numbers = _extract_all_numbers_custom(result_str)
                
                if numbers:
                    step_key = f"步骤{step_num}"
                    loop_key = f"循环{loop_index}"
                    
                    if step_key not in chart_data:
                        chart_data[step_key] = {
                            "step_num": int(step_num),
                            "content": step_content.replace('?', '').strip(),
                            "loops": {},
                            "loop_colors": {}
                        }
                    
                    if loop_key not in chart_data[step_key]["loops"]:
                        chart_data[step_key]["loops"][loop_key] = {
                            "loop_index": loop_index,
                            "data_points": numbers
                        }
                        
                        # 为每个循环分配颜色
                        color_map = plt.cm.tab10(np.linspace(0, 1, 10))
                        color_idx = (loop_index - 1) % len(color_map)
                        chart_data[step_key]["loop_colors"][loop_key] = color_map[color_idx]
    
    return chart_data

def _extract_all_numbers_custom(text: str) -> List[float]:
    """从文本中提取所有数字（自定义版本）"""
    try:
        if not text or not isinstance(text, str):
            return []
        
        clean_text = str(text).strip()
        
        # 检查错误信息
        error_keywords = ["仪器通信错误", "VI_ERROR_TMO", "Timeout", "通信失败", "错误", "ERROR", "失败"]
        if any(keyword in clean_text.upper() for keyword in [k.upper() for k in error_keywords]):
            return []
        
        if clean_text in ['', 'N/A', 'None', 'null', '[]', '{}', '命令执行成功']:
            return []
        
        # 处理常见的返回格式
        # 1. 逗号分隔的数字
        if ',' in clean_text:
            parts = [part.strip() for part in clean_text.split(',')]
            numbers = []
            for part in parts:
                # 尝试直接转换
                try:
                    num = float(part)
                    numbers.append(num)
                except ValueError:
                    # 如果失败，尝试从字符串中提取数字
                    num_match = re.search(r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?', part)
                    if num_match:
                        try:
                            numbers.append(float(num_match.group()))
                        except:
                            continue
            if numbers:
                return numbers
        
        # 2. 单个数字
        try:
            num = float(clean_text)
            return [num]
        except ValueError:
            pass
        
        # 3. 使用正则表达式提取所有数字
        pattern = r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?'
        matches = re.findall(pattern, clean_text)
        
        numbers = []
        for match in matches:
            try:
                number = float(match)
                numbers.append(number)
            except ValueError:
                continue
        
        # 4. 如果正则提取失败，尝试更宽松的匹配
        if not numbers and clean_text:
            # 查找任何看起来像数字的字符串
            num_candidates = re.findall(r'\b\d+\.?\d*\b', clean_text)
            for candidate in num_candidates:
                try:
                    numbers.append(float(candidate))
                except:
                    continue
        
        return numbers
    except Exception as e:
        return []

# 在 souren_exporter.py 中修改数据处理部分，确保每次循环的数据都被正确提取

def extract_chart_data_from_execution_details(execution_details: List[Dict], loop_count: int = 1) -> Dict:
    """从执行详情中提取图表数据，支持多次循环"""
    from souren_config import CHART_DATA_EXTRACTION_CONFIG, USE_CHART_DATA_EXTRACTION
    
    if not USE_CHART_DATA_EXTRACTION:
        return {}
    
    chart_data = {}
    
    # 按循环次数组织数据
    for loop_idx in range(1, loop_count + 1):
        loop_data = {}
        
        # 过滤出当前循环的数据
        loop_details = [d for d in execution_details if d.get("loop_iteration") == loop_idx]
        
        # 提取数据
        for detail in loop_details:
            command = detail.get("content", "")
            result = detail.get("result", "")
            
            # 检查是否需要提取此命令的数据
            for cmd_pattern, indices in CHART_DATA_EXTRACTION_CONFIG.items():
                if cmd_pattern in command:
                    # 清理结果字符串
                    result_clean = str(result).strip().strip('"').strip("'")
                    
                    # 分割结果
                    parts = result_clean.split(',')
                    
                    # 提取指定位置的数据
                    for idx in indices:
                        if 0 < idx <= len(parts):
                            value_str = parts[idx-1].strip()
                            try:
                                value = float(value_str)
                            except ValueError:
                                # 如果不是数字，跳过
                                continue
                            
                            # 创建数据键
                            data_key = f"{cmd_pattern}_idx{idx}"
                            if data_key not in loop_data:
                                loop_data[data_key] = []
                            
                            # 添加数据
                            loop_data[data_key].append({
                                "step": detail.get("step", 0),
                                "value": value,
                                "command": command,
                                "loop_iteration": loop_idx
                            })
        
        if loop_data:
            chart_data[f"循环{loop_idx}"] = loop_data
    
    return chart_data

if __name__ == "__main__":
    exporter = ResultExporter()
    success = exporter.convert_to_excel()
    
    if success:
        print("✅ Excel文件导出成功!")
    else:
        print("❌ Excel文件导出失败")