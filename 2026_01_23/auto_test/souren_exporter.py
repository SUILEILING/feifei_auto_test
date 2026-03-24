from lib.var import *
import pandas as pd
import json
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import numpy as np
import re
from typing import List, Dict, Optional
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

class ResultExporter:
    """结果导出器 - 将JSON结果自动转换为Excel,使用matplotlib绘制柱状图"""
    
    def __init__(self):
        # 延迟导入，避免循环导入问题
        import souren_config
        
        # 获取配置
        self.log_dir = souren_config.EXECUTION_DIR
        
        # 应用配置值（使用默认值避免缺失配置）
        self.default_row_height = getattr(souren_config, 'EXCEL_DEFAULT_ROW_HEIGHT', 13.5)
        self.default_column_width = getattr(souren_config, 'EXCEL_DEFAULT_COLUMN_WIDTH', 9)
        self.header_row_height = getattr(souren_config, 'EXCEL_HEADER_ROW_HEIGHT', 20)
        
        # 应用列宽配置（使用安全获取方式）
        self.summary_column_widths = getattr(souren_config, 'EXCEL_SUMMARY_COLUMN_WIDTHS', {})
        self.details_column_widths = getattr(souren_config, 'EXCEL_DETAILS_COLUMN_WIDTHS', {})
        
        # 图表列宽配置（如果不存在则使用默认值）
        if hasattr(souren_config, 'EXCEL_CHART_COLUMN_WIDTHS'):
            self.chart_column_widths = souren_config.EXCEL_CHART_COLUMN_WIDTHS
        else:
            self.chart_column_widths = {'default': 15}
            print(f"⚠️  配置中缺少EXCEL_CHART_COLUMN_WIDTHS，使用默认值")
        
        # 应用颜色配置
        self.border_colors = getattr(souren_config, 'EXCEL_BORDER_COLORS', {"default": "000000"})
        
        # 生成图表颜色的函数（如果存在）
        if hasattr(souren_config, 'generate_chart_colors'):
            self.generate_chart_colors_func = souren_config.generate_chart_colors
        else:
            # 默认颜色生成函数
            self.generate_chart_colors_func = lambda count: plt.cm.tab10(np.linspace(0, 1, min(count, 10)))
        
        # 获取图表配置开关
        self.use_chart_data_extraction = getattr(souren_config, 'USE_CHART_DATA_EXTRACTION', False)
        
        # 根据开关选择配置方式
        if self.use_chart_data_extraction:
            # 使用新的命令模式提取配置
            self.data_extraction_config = getattr(souren_config, 'CHART_DATA_EXTRACTION_CONFIG', {})
            self.chart_steps_filter = []  # 不使用步骤编号筛选
            
            print(f"📊 使用命令模式提取图表数据")
            print(f"   配置内容: {self.data_extraction_config}")
        else:
            # 使用旧的步骤编号筛选配置
            self.data_extraction_config = {}
            self.chart_steps_filter = getattr(souren_config, 'CHART_STEPS_FILTER', [])
            print(f"📊 使用步骤编号筛选配置")
            print(f"   配置内容: {self.chart_steps_filter}")
        
        # 创建样式对象（使用安全获取方式）
        self.header_fill = PatternFill(
            start_color=getattr(souren_config, 'EXCEL_HEADER_FILL_COLOR', "366092"), 
            end_color=getattr(souren_config, 'EXCEL_HEADER_FILL_COLOR', "366092"), 
            fill_type="solid"
        )
        self.header_font = Font(
            color=getattr(souren_config, 'EXCEL_HEADER_FONT_COLOR', "FFFFFF"), 
            bold=True, 
            size=getattr(souren_config, 'EXCEL_HEADER_FONT_SIZE', 11)
        )
        self.query_result_fill = PatternFill(
            start_color=getattr(souren_config, 'EXCEL_QUERY_RESULT_FILL', "E2F0D9"), 
            end_color=getattr(souren_config, 'EXCEL_QUERY_RESULT_FILL', "E2F0D9"), 
            fill_type="solid"
        )
        self.query_result_font = Font(
            color=getattr(souren_config, 'EXCEL_QUERY_RESULT_FONT', "006400"), 
            bold=True
        )
        self.data_font = Font(size=getattr(souren_config, 'EXCEL_DATA_FONT_SIZE', 10))
        
        # 对齐样式
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # 图表样式配置
        self.chart_style = {
            'figsize': (14, 8),
            'dpi': 150,
            'grid_alpha': 0.3,
            'font_size_title': 14,
            'font_size_axis': 12,
            'legend_loc': 'best',
            'bar_width': 0.6,
            'bar_alpha': 0.8,
            'colors': plt.cm.Set3(np.linspace(0, 1, 12))  # 使用Set3颜色映射，支持多个循环
        }
    
    def find_latest_json_result(self) -> Optional[str]:
        """查找最新的JSON结果文件"""
        try:
            if not os.path.exists(self.log_dir):
                return None
            
            json_files = []
            for filename in os.listdir(self.log_dir):
                if filename.startswith("souren_results_") and filename.endswith(".json"):
                    filepath = os.path.join(self.log_dir, filename)
                    if os.path.isfile(filepath):
                        mtime = os.path.getmtime(filepath)
                        json_files.append({"path": filepath, "mtime": mtime})
            
            if not json_files:
                return None
            
            json_files.sort(key=lambda x: x["mtime"], reverse=True)
            return json_files[0]["path"]
            
        except Exception:
            return None
    
    def load_json_results(self, json_file: str) -> List[Dict]:
        """加载JSON结果数据"""
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            if not content.strip():
                return []
            
            results = json.loads(content)
            return [results] if isinstance(results, dict) else results
            
        except Exception:
            return []
    
    def convert_to_excel(self, json_file: str = None) -> bool:
        """将JSON结果转换为Excel文件"""
        try:
            if json_file is None:
                json_file = self.find_latest_json_result()
                if json_file is None:
                    print("❌ 未找到JSON结果文件")
                    return False
            
            json_name = os.path.basename(json_file)
            excel_name = json_name.replace('.json', '.xlsx')
            excel_file = os.path.join(os.path.dirname(json_file), excel_name)
            
            # 检查文件是否被锁定，如果是则创建新文件名
            import time
            if os.path.exists(excel_file):
                try:
                    # 尝试重命名文件
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                    backup_file = excel_file.replace('.xlsx', f'_{timestamp}.xlsx')
                    os.rename(excel_file, backup_file)
                    print(f"⚠️ 原文件已存在，已重命名为: {os.path.basename(backup_file)}")
                except:
                    # 如果重命名失败，使用新文件名
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    excel_name = f"souren_results_{timestamp}.xlsx"
                    excel_file = os.path.join(os.path.dirname(json_file), excel_name)
                    print(f"⚠️ 文件被锁定，使用新文件名: {excel_name}")
            
            results = self.load_json_results(json_file)
            if not results:
                print("❌ 未加载到有效的JSON结果数据")
                return False
            
            print(f"📊 开始创建Excel文件: {excel_name}")
            
            # 创建Excel文件
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                self._create_summary_sheet(writer, results)
                self._create_execution_details_sheet(writer, results)
            
            # 应用样式和图表
            self._apply_styles_and_charts(excel_file, results)
            
            print(f"✅ Excel文件已生成: {excel_file}")
            return True
            
        except Exception as e:
            print(f"❌ Excel文件导出失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter, results: List[Dict]):
        """创建汇总表"""
        try:
            valid_results = [r for r in results if r.get('executed_steps', 0) > 0]
            
            if not valid_results:
                df_empty = pd.DataFrame({"状态": ["没有有效的执行记录"]})
                df_empty.to_excel(writer, sheet_name='执行汇总', index=False)
                return
            
            summary_data = []
            for i, result in enumerate(valid_results):
                executed = result.get('executed_steps', 0)
                passed = result.get('passed', 0)
                success_rate = round((passed / executed) * 100, 2) if executed > 0 else 0
                
                summary = {
                    "序号": i + 1,
                    "执行时间": result.get('timestamp_readable', result.get('timestamp', 'N/A')),
                    "SCV文件": result.get('file', 'N/A'),
                    "设备": self._extract_device_name(result),
                    "执行模式": result.get('mode', 'run_all'),
                    "循环次数": result.get('loop_count', 1),
                    "总步骤数": result.get('total_steps', result.get('total_executions', 0)),
                    "已执行步骤": executed,
                    "通过步骤": passed,
                    "失败步骤": result.get('failed', 0),
                    "成功率(%)": success_rate,
                    "总耗时(秒)": round(float(result.get('execution_time', result.get('duration', 0))), 2),
                    "状态": result.get('status', 'N/A'),
                    "状态消息": result.get('message', 'N/A')
                }
                summary_data.append(summary)
            
            if summary_data:
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='执行汇总', index=False)
            
        except Exception as e:
            print(f"创建汇总表失败: {e}")
    
    def _extract_device_name(self, result: Dict) -> str:
        """提取设备名称"""
        device = result.get('device', {})
        if isinstance(device, dict):
            return device.get('name', device.get('address', 'N/A'))
        elif isinstance(device, str):
            return device
        else:
            return 'N/A'
    
    def _create_execution_details_sheet(self, writer: pd.ExcelWriter, results: List[Dict]):
        """创建详细执行记录表"""
        try:
            details_data = []
            
            for result_index, result in enumerate(results):
                if result.get('executed_steps', 0) == 0:
                    continue
                    
                execution_details = result.get('execution_details', [])
                loop_count = result.get('loop_count', 1)
                
                if result.get('mode', 'run_all') != 'loop_info' or loop_count <= 1:
                    for detail in execution_details:
                        details_data.append(self._create_detail_record(detail, result_index, result, 1, 1))
                    continue
                
                # 按循环轮次分组
                loop_groups = {}
                for detail in execution_details:
                    loop_index = detail.get('loop_index', 1)
                    if loop_index not in loop_groups:
                        loop_groups[loop_index] = []
                    loop_groups[loop_index].append(detail)
                
                for loop_index in sorted(loop_groups.keys()):
                    if loop_index > 1:
                        details_data.extend([{}, {}])
                    for detail in loop_groups[loop_index]:
                        details_data.append(self._create_detail_record(
                            detail, result_index, result, loop_index, loop_count
                        ))
            
            if details_data:
                columns = [
                    "汇总序号", "执行时间", "步骤序号", "循环轮次", "总循环数", 
                    "步骤内容", "步骤类型", "执行状态", "执行结果", "耗时(秒)"
                ]
                
                formatted_data = []
                for record in details_data:
                    formatted_record = {}
                    for col in columns:
                        formatted_record[col] = record.get(col, "")
                    formatted_data.append(formatted_record)
                
                df_details = pd.DataFrame(formatted_data, columns=columns)
                df_details.to_excel(writer, sheet_name='详细执行记录', index=False)
            
        except Exception as e:
            print(f"创建详细记录表失败: {e}")
    
    def _create_detail_record(self, detail: Dict, result_index: int, result: Dict, 
                            loop_index: int, total_loops: int) -> Dict:
        """创建单个详细记录"""
        return {
            "汇总序号": result_index + 1,
            "执行时间": result.get('timestamp_readable', result.get('timestamp', 'N/A')),
            "步骤序号": detail.get('original_step', detail.get('step', 'N/A')),
            "循环轮次": loop_index,
            "总循环数": total_loops,
            "步骤内容": detail.get('content', ''),
            "步骤类型": detail.get('type', 'Normal'),
            "执行状态": detail.get('status', 'N/A'),
            "执行结果": str(detail.get('result', 'N/A')),
            "耗时(秒)": round(float(detail.get('duration', 0)), 3)
        }
    
    def _extract_chart_data(self, results: List[Dict]) -> Dict:
        """提取图表数据"""
        if self.use_chart_data_extraction:
            return self._extract_chart_data_by_command(results)
        else:
            return self._extract_chart_data_by_step(results)
    
    def _extract_chart_data_by_command(self, results: List[Dict]) -> Dict:
        """根据配置的命令模式提取图表数据"""
        chart_data = {}
        
        print(f"📊 开始根据命令模式提取数据...")
        print(f"   配置的命令模式: {list(self.data_extraction_config.keys())}")
        
        for result in results:
            execution_details = result.get('execution_details', [])
            loop_count = result.get('loop_count', 1)
            
            for detail in execution_details:
                step_content = detail.get('content', '')
                result_str = str(detail.get('result', ''))
                step_num = detail.get('original_step', detail.get('step', 0))
                loop_index = detail.get('loop_index', 1)
                
                # 检查是否匹配配置的命令模式
                for command_pattern, positions in self.data_extraction_config.items():
                    if command_pattern in step_content:
                        # 提取数字
                        numbers = self._extract_all_numbers(result_str)
                        
                        if numbers:
                            # 为每个提取位置创建数据系列
                            for pos_idx, position in enumerate(positions):
                                if 1 <= position <= len(numbers):
                                    value = numbers[position - 1]
                                    
                                    # 生成数据系列键
                                    series_key = f"{command_pattern}_位置{position}"
                                    
                                    if series_key not in chart_data:
                                        chart_data[series_key] = {
                                            "command": command_pattern,
                                            "position": position,
                                            "data": {},
                                            "loop_indices": set()
                                        }
                                    
                                    # 存储循环数据
                                    loop_key = f"循环{loop_index}"
                                    if loop_key not in chart_data[series_key]["data"]:
                                        chart_data[series_key]["data"][loop_key] = value
                                        chart_data[series_key]["loop_indices"].add(loop_index)
                                    
                                    print(f"    ✅ 提取到数据: {command_pattern} 循环{loop_index} 位置{position} = {value:.3f}")
        
        print(f"📊 共提取了 {len(chart_data)} 个数据系列")
        return chart_data
    
    def _extract_chart_data_by_step(self, results: List[Dict]) -> Dict:
        """按步骤编号提取图表数据"""
        chart_data = {}
        
        print(f"📊 使用步骤编号筛选数据...")
        print(f"   筛选步骤: {self.chart_steps_filter}")
        
        for result in results:
            execution_details = result.get('execution_details', [])
            
            for detail in execution_details:
                step_content = detail.get('content', '')
                result_str = str(detail.get('result', ''))
                step_num = detail.get('original_step', detail.get('step', 0))
                loop_index = detail.get('loop_index', 1)
                
                # 根据配置筛选步骤
                if step_num not in self.chart_steps_filter:
                    continue
                
                # 只处理查询命令且有结果的数据
                if '?' in step_content and result_str and result_str != 'N/A':
                    numbers = self._extract_all_numbers(result_str)
                    
                    if numbers:
                        step_key = f"步骤{step_num}"
                        loop_key = f"循环{loop_index}"
                        
                        if step_key not in chart_data:
                            chart_data[step_key] = {
                                "step_num": int(step_num),
                                "content": step_content.replace('?', '').strip(),
                                "loops": {}
                            }
                        
                        # 存储该循环的数据点
                        if loop_key not in chart_data[step_key]["loops"]:
                            chart_data[step_key]["loops"][loop_key] = {
                                "loop_index": loop_index,
                                "data_points": numbers
                            }
        
        print(f"📊 根据步骤编号筛选了 {len(chart_data)} 个步骤的数据")
        return chart_data
    
    def _extract_all_numbers(self, text: str) -> List[float]:
        """从文本中提取所有数字，保持原值"""
        try:
            if not text:
                return []
            
            clean_text = text.strip()
            
            # 检查是否包含错误信息
            error_keywords = ["仪器通信错误", "VI_ERROR_TMO", "Timeout", "通信失败", "错误", "ERROR"]
            if any(keyword in clean_text for keyword in error_keywords):
                return []
            
            # 检查是否为空或无效响应
            if clean_text in ['', 'N/A', 'None', 'null', '[]', '{}']:
                return []
            
            # 处理逗号分隔的数组格式（这是最常见的情况）
            if ',' in clean_text:
                parts = [part.strip() for part in clean_text.split(',')]
                numbers = []
                for part in parts:
                    try:
                        # 尝试直接转换为数字
                        num = float(part)
                        numbers.append(num)
                    except ValueError:
                        # 如果不是纯数字，尝试提取数字
                        pattern = r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?'
                        matches = re.findall(pattern, part)
                        for match in matches:
                            try:
                                numbers.append(float(match))
                            except ValueError:
                                continue
                if numbers:
                    return numbers
            
            # 尝试直接解析为单个数字
            try:
                num = float(clean_text)
                return [num]
            except ValueError:
                pass
            
            # 常规提取所有数字
            pattern = r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?'
            matches = re.findall(pattern, clean_text)
            
            numbers = []
            for match in matches:
                try:
                    number = float(match)
                    numbers.append(number)
                except ValueError:
                    continue
            
            return numbers
        except Exception as e:
            print(f"❌ 提取数字失败: {e}, 文本: {text[:100]}...")
            return []
    
    def _create_bar_chart(self, command: str, position: int, 
                         loop_data: Dict[str, float], loop_indices: set) -> BytesIO:
        """
        创建柱状图显示不同循环的数据
        """
        try:
            # 准备数据
            sorted_loop_keys = sorted(loop_data.keys(), key=lambda x: int(x.replace('循环', '')))
            loop_values = []
            loop_labels = []
            
            for loop_key in sorted_loop_keys:
                loop_labels.append(loop_key)
                loop_values.append(loop_data[loop_key])
            
            if not loop_values:
                return None
            
            # 创建图表
            fig, ax = plt.subplots(figsize=self.chart_style['figsize'])
            
            # 为每个循环使用不同的颜色
            colors = self.chart_style['colors'][:len(loop_values)]
            
            # 创建柱状图
            bars = ax.bar(range(len(loop_values)), loop_values, 
                         color=colors, 
                         alpha=self.chart_style['bar_alpha'],
                         width=self.chart_style['bar_width'])
            
            # 添加数值标签
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{height:.3f}',
                       ha='center', va='bottom',
                       fontsize=10)
            
            # 设置图表属性
            ax.set_xlabel('循环轮次', fontsize=self.chart_style['font_size_axis'])
            ax.set_ylabel('数值', fontsize=self.chart_style['font_size_axis'])
            
            # 提取命令名称
            command_name = self._get_command_name(command)
            title = f'{command_name} (位置{position})'
            ax.set_title(title, fontsize=self.chart_style['font_size_title'])
            
            # 设置x轴刻度
            ax.set_xticks(range(len(loop_labels)))
            ax.set_xticklabels(loop_labels)
            
            # 添加网格
            ax.grid(True, alpha=self.chart_style['grid_alpha'], axis='y')
            
            # 调整布局
            plt.tight_layout()
            
            # 将图表保存到BytesIO
            img_data = BytesIO()
            plt.savefig(img_data, format='png', dpi=self.chart_style['dpi'], 
                       bbox_inches='tight', facecolor='white')
            img_data.seek(0)
            plt.close(fig)  # 关闭图形释放内存
            
            return img_data
            
        except Exception as e:
            print(f"❌ 创建柱状图失败: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _get_command_name(self, command: str) -> str:
        """从命令中提取简化的名称"""
        # 移除查询符号
        clean_command = command.replace('?', '').strip()
        
        # 提取关键部分
        if 'BLER:UL' in clean_command or 'UL:RESult' in clean_command:
            return 'UL BLER'
        elif 'BLER:DL' in clean_command or 'DL:RESult' in clean_command:
            return 'DL BLER'
        elif 'TXP:AVG' in clean_command:
            return 'TXP AVG'
        elif 'TXP:MIN' in clean_command:
            return 'TXP MIN'
        elif 'TXP:MAX' in clean_command:
            return 'TXP MAX'
        
        # 提取最后一部分作为名称
        parts = clean_command.split(':')
        if parts:
            return parts[-1]
        
        return clean_command[:20]
    
    def _apply_styles_and_charts(self, excel_file: str, results: List[Dict]):
        """应用样式和图表到Excel文件"""
        try:
            wb = load_workbook(excel_file)
            
            # 1. 格式化执行汇总表
            if '执行汇总' in wb.sheetnames:
                self._format_summary_sheet(wb['执行汇总'])
            
            # 2. 格式化详细执行记录表
            if '详细执行记录' in wb.sheetnames:
                self._format_details_sheet(wb['详细执行记录'])
            
            # 3. 提取图表数据
            chart_data = self._extract_chart_data(results)
            
            if chart_data:
                # 4. 创建数据分析图表工作表
                self._create_chart_sheet(wb, chart_data)
            else:
                # 创建空的数据分析图表工作表
                if '数据分析图表' in wb.sheetnames:
                    wb.remove(wb['数据分析图表'])
                chart_ws = wb.create_sheet(title='数据分析图表')
                chart_ws['A1'] = "数据分析图表"
                
                if self.use_chart_data_extraction:
                    chart_ws['A2'] = "⚠️ 没有找到配置命令的查询数据"
                    chart_ws['A3'] = f"配置的命令模式: {list(self.data_extraction_config.keys())}"
                    print(f"⚠️ 没有找到配置命令的查询数据")
                    print(f"   配置的命令模式: {list(self.data_extraction_config.keys())}")
                else:
                    chart_ws['A2'] = "⚠️ 没有找到配置步骤的查询数据"
                    chart_ws['A3'] = f"配置的步骤编号: {self.chart_steps_filter}"
                    print(f"⚠️ 没有找到配置步骤的查询数据")
            
            # 5. 调整列宽行高
            self._adjust_dimensions(wb)
            
            wb.save(excel_file)
            
        except Exception as e:
            print(f"❌ 应用样式和图表失败: {e}")
            import traceback
            traceback.print_exc()
    
    def _create_chart_sheet(self, wb, chart_data):
        """创建图表工作表"""
        try:
            # 移除旧的图表工作表
            if '数据分析图表' in wb.sheetnames:
                wb.remove(wb['数据分析图表'])
            
            # 创建新的数据分析图表工作表
            chart_ws = wb.create_sheet(title='数据分析图表')
            
            # 1. 添加标题
            title = "数据分析图表（基于命令模式提取）"
            config_info = f"数据提取配置: {self.data_extraction_config}"
            
            chart_ws['A1'] = title
            chart_ws['A1'].font = Font(size=14, bold=True, color=self.header_fill.start_color.rgb)
            chart_ws.merge_cells('A1:L1')
            
            # 显示配置信息
            chart_ws['A2'] = config_info
            chart_ws['A2'].font = Font(size=10, italic=True)
            
            print(f"📊 开始创建柱状图...")
            
            # 2. 为每个数据系列创建柱状图
            current_row = 5
            charts_created = 0
            
            for series_key, series_info in chart_data.items():
                command = series_info["command"]
                position = series_info["position"]
                loop_data = series_info["data"]
                
                if not loop_data:
                    continue
                
                print(f"  📊 处理命令: {command}, 位置: {position}")
                print(f"    循环数据: {loop_data}")
                
                # 创建柱状图
                img_data = self._create_bar_chart(command, position, loop_data, series_info["loop_indices"])
                
                if img_data:
                    # 创建Excel图像对象
                    img = ExcelImage(img_data)
                    
                    # 调整图像大小
                    img.width = 800  # 像素宽度
                    img.height = 400  # 像素高度
                    
                    # 将图像插入到Excel中
                    chart_cell = f"A{current_row}"
                    chart_ws.add_image(img, chart_cell)
                    
                    # 在图表下方添加简要说明
                    explanation_row = current_row + 20
                    chart_ws.cell(row=explanation_row, column=1, value=f"命令: {command}")
                    chart_ws.cell(row=explanation_row + 1, column=1, value=f"提取位置: {position}")
                    chart_ws.cell(row=explanation_row + 2, column=1, value=f"循环数: {len(loop_data)}")
                    
                    # 显示具体数值
                    chart_ws.cell(row=explanation_row + 3, column=1, value="具体数值:")
                    value_row = explanation_row + 4
                    for loop_key, value in sorted(loop_data.items()):
                        chart_ws.cell(row=value_row, column=1, value=f"{loop_key}: {value:.3f}")
                        value_row += 1
                    
                    # 更新下一张图表的起始位置
                    current_row = value_row + 3
                    charts_created += 1
                else:
                    print(f"  ❌ 创建图表失败: {command}")
            
            if charts_created == 0:
                chart_ws['A5'] = "⚠️ 没有足够的数据来创建图表"
                chart_ws['A6'] = "可能原因:"
                chart_ws['A7'] = "1. 数据点数量不足"
                chart_ws['A8'] = "2. 查询命令没有返回有效的数字数据"
                chart_ws['A9'] = f"3. 配置的命令模式不匹配: {list(self.data_extraction_config.keys())}"
                
                print("⚠️ 没有足够的数据来创建图表")
            else:
                print(f"✅ 成功创建了 {charts_created} 个柱状图并插入到Excel")
            
            # 3. 设置列宽
            for col_letter, width in self.chart_column_widths.items():
                if col_letter == 'default':
                    # 设置默认列宽
                    for col in range(1, chart_ws.max_column + 1):
                        col_letter = get_column_letter(col)
                        if col_letter not in self.chart_column_widths:
                            chart_ws.column_dimensions[col_letter].width = width
                else:
                    if col_letter in chart_ws.column_dimensions:
                        chart_ws.column_dimensions[col_letter].width = width
            
        except Exception as e:
            print(f"❌ 创建图表工作表失败: {e}")
            import traceback
            traceback.print_exc()
    
    def _format_summary_sheet(self, ws):
        """格式化汇总表"""
        try:
            # 格式化表头
            for cell in ws[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.border = self._create_border(self.border_colors.get("default", "000000"))
            
            # 格式化数据行
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = self._create_border(self.border_colors.get("default", "000000"))
                    cell.font = self.data_font
                    
                    # 设置对齐方式
                    col_idx = cell.column
                    if col_idx in [1, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
                        cell.alignment = self.center_align
                    elif col_idx == 2:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = self.left_align
                
                # 设置行高
                ws.row_dimensions[row[0].row].height = self.default_row_height
            
        except Exception as e:
            print(f"❌ 格式化汇总表失败: {e}")
    
    def _format_details_sheet(self, ws):
        """格式化详细执行记录表"""
        try:
            # 格式化表头
            for cell in ws[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_align
                cell.border = self._create_border(self.border_colors.get("default", "000000"))
            
            # 设置表头行高
            ws.row_dimensions[1].height = self.header_row_height
            
            # 处理数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                if not any(cell.value for cell in row):
                    continue
                
                # 检查是否是查询命令
                step_content_cell = row[5]  # F列是步骤内容
                is_query = step_content_cell.value and '?' in str(step_content_cell.value)
                
                for cell in row:
                    cell.border = self._create_border(self.border_colors.get("default", "000000"))
                    cell.font = self.data_font
                    
                    # 设置对齐方式
                    col_idx = cell.column
                    if col_idx in [1, 3, 4, 5, 7, 8, 10]:
                        cell.alignment = self.center_align
                    elif col_idx == 2:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = self.left_align
                
                # 高亮显示查询结果行
                if is_query:
                    for cell in row:
                        cell.fill = self.query_result_fill
                        cell.font = self.query_result_font
                
                # 设置数据行高
                ws.row_dimensions[row_idx].height = self.default_row_height
            
        except Exception as e:
            print(f"❌ 格式化详细记录表失败: {e}")
    
    def _create_border(self, color: str) -> Border:
        """创建边框"""
        side = Side(style='thin', color=color)
        return Border(left=side, right=side, top=side, bottom=side)
    
    def _adjust_dimensions(self, wb):
        """调整列宽和行高"""
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 设置行高
                for row in range(1, ws.max_row + 1):
                    if row == 1:
                        # 表头行高
                        ws.row_dimensions[row].height = self.header_row_height
                    else:
                        # 数据行高
                        ws.row_dimensions[row].height = self.default_row_height
                
                # 根据工作表类型应用不同的列宽设置
                if sheet_name == '执行汇总':
                    # 首先设置默认列宽
                    for col in ws.columns:
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = self.default_column_width
                    
                    # 然后应用特殊列宽配置
                    for col_letter, width in self.summary_column_widths.items():
                        if col_letter in ws.column_dimensions:
                            ws.column_dimensions[col_letter].width = width
                            
                elif sheet_name == '详细执行记录':
                    # 首先设置默认列宽
                    for col in ws.columns:
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = self.default_column_width
                    
                    # 然后应用特殊列宽配置
                    for col_letter, width in self.details_column_widths.items():
                        if col_letter in ws.column_dimensions:
                            ws.column_dimensions[col_letter].width = width
                            
                elif sheet_name == '数据分析图表':
                    # 应用图表列宽配置
                    for col_letter, width in self.chart_column_widths.items():
                        if col_letter == 'default':
                            # 设置默认列宽
                            for col in range(1, ws.max_column + 1):
                                col_letter = get_column_letter(col)
                                if col_letter not in self.chart_column_widths:
                                    ws.column_dimensions[col_letter].width = width
                        elif col_letter in ws.column_dimensions:
                            ws.column_dimensions[col_letter].width = width
                            
                else:
                    # 其他工作表设置默认列宽
                    for col in ws.columns:
                        col_letter = get_column_letter(col[0].column)
                        ws.column_dimensions[col_letter].width = self.default_column_width
                
                # 启用筛选（除了图表工作表）
                if sheet_name != '数据分析图表' and ws.max_row > 1:
                    try:
                        ws.auto_filter.ref = ws.dimensions
                    except:
                        pass
            
            print(f"✅ 已应用所有Excel导出配置")
            print(f"   - 当前模式: {'命令模式提取' if self.use_chart_data_extraction else '步骤编号筛选'}")
            print(f"   - 图表类型: 柱状图（每个循环不同颜色）")
            
        except Exception as e:
            print(f"❌ 调整列宽行高失败: {e}")

# 保留独立的导出功能
if __name__ == "__main__":
    exporter = ResultExporter()
    
    # 可以传入特定的json文件路径
    # x = r"D:\feifei\2026_01_05\auto_test\log\execution_20260113_163555\souren_results_20260113_163555.json"
    success = exporter.convert_to_excel()
    
    if success:
        print("✅ Excel文件导出成功!")
    else:
        print("❌ Excel文件导出失败")